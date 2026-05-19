import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def create_template(output_path):
    wb = Workbook()
    
    # --- Sheet 1: Setup ---
    ws_setup = wb.active
    ws_setup.title = "Setup"
    ws_setup.append(["ID", "Algorithm", "Hyper-Parameter Settings", "Note"])
    ws_setup.append(["1", "Logistic Regression (LR)", "max_iter=1000, solver='lbfgs', multi_class='auto'", "Mốc chuẩn tuyến tính"])
    ws_setup.append(["2", "Linear SVC (SVM)", "dual=False, C=1.0, penalty='l2', max_iter=1000", "Phân tách không gian"])
    ws_setup.append(["3", "Random Forest (RF)", "n_estimators=100, max_depth=10, min_samples_split=2, min_samples_leaf=1", "Bagging Tree"])
    ws_setup.append(["4", "XGBoost (XGB)", "n_estimators=100, max_depth=6, learning_rate=0.3, objective='multi:softprob'", "Tree Boosting"])
    ws_setup.append(["5", "LightGBM (LGBM)", "n_estimators=100, max_depth=6, num_leaves=31, learning_rate=0.1", "Fast Tree Boosting"])
    ws_setup.append(["6", "CatBoost (CAT)", "iterations=100, depth=6, learning_rate=0.03", "Categorical Boosting"])
    ws_setup.append(["7", "TabNet", "n_d=8, n_a=8, n_steps=3, gamma=1.3, momentum=0.02", "Deep Learning for Tabular"])
    
    # --- Sheet 2: Scenario ---
    ws_scen = wb.create_sheet("Scenario")
    ws_scen.append(["Scenario ID", "Dataset", "Coreset Method", "Budget Ratio", "Seeds", "Evaluators"])
    ws_scen.append(["1", "CourseQuality", "13 SOTA Methods", "1%, 5%, 10%, 20%", "10 Seeds (0-9)", "LR, SVM, RF, XGB, LGBM, CAT, TABNET"])
    
    # --- Sheet 3: Data Course_Quality ---
    ws_data = wb.create_sheet("Data Course_Quality")
    ws_data.append(["ID", "Feature Name", "Type", "Missing Values", "Unique Values", "Description", "Notes"])
    
    # --- Sheet 4: Architecture ---
    ws_arch = wb.create_sheet("Architecture")
    ws_arch.append(["Component", "Path", "Description", "Role in Benchmark"])
    ws_arch.append(["Giai đoạn 1", "src/data_station.py", "Xử lý dữ liệu", "Chuẩn hóa, LabelEncode, Nền tảng"])
    ws_arch.append(["Giai đoạn 2", "src/coreset/factory.py", "Sinh Coreset (Kaggle)", "Chạy 13 thuật toán, thu thập dữ liệu nén"])
    ws_arch.append(["Giai đoạn 3", "src/training/train_arena.py", "Chấm điểm (Local)", "Đo lường bằng XGBoost, LightGBM, Random Forest..."])
    
    # --- Sheet 5: Course Quality (Results) ---
    ws_res = wb.create_sheet("Course Quality")
    
    # Row 1
    row1 = ["Baseline", "Model", "Phase", "Data Quality", "", "", "", "", "", "", "", "", "", "", "", "Time", "", "Performance"]
    row1.extend([""] * 22)
    ws_res.append(row1)
    
    # Row 2
    row2 = ["", "", "", "Completeness", "Consistency", "snan", "smaj", "sjsd", "sent", "sdrift", "S_eff", "sleak", "S_san+", "Sperf", "Acc-DQ", 
            "Build Model", "Predict", 
            "Accuracy", "BalancedAcc", "Precision Macro", "Precision Weighted", "Recall Macro", "Recall Weighted", "F1-Score Macro", "F1-Score Weighted", "GMean", "MCC", "Kappa",
            "Label (Exellent)", "", "", "", "Label (Good)", "", "", "", "Label (Average)", "", "", ""]
    ws_res.append(row2)
    
    # Row 3
    row3 = [""] * 28
    row3.extend(["Precision", "Recall", "F1-Score", "G-Mean"] * 3)
    ws_res.append(row3)
    
    # Merging cells in Course Quality
    ws_res.merge_cells('D1:O1') # Data Quality
    ws_res.merge_cells('P1:Q1') # Time
    ws_res.merge_cells('R1:AN1') # Performance
    
    ws_res.merge_cells('A1:A3') # Baseline
    ws_res.merge_cells('B1:B3') # Model
    ws_res.merge_cells('C1:C3') # Phase
    
    for col in range(4, 29): # Completeness to Kappa
        ws_res.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
        
    ws_res.merge_cells('AC2:AF2') # Label Exellent
    ws_res.merge_cells('AG2:AJ2') # Label Good
    ws_res.merge_cells('AK2:AN2') # Label Average
    
    # Formatting
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(min_row=1, max_row=3):
            for cell in row:
                cell.font = header_font
                cell.alignment = center_align
                if cell.value:
                    cell.fill = header_fill
                    
        # Adjust column widths
        for col in sheet.columns:
            max_length = 0
            column_idx = col[0].column
            column_letter = get_column_letter(column_idx)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = min(adjusted_width, 25)
            
    # Save
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Generated template: {output_path}")

if __name__ == "__main__":
    out_path = r"C:\KLTN\paper\Result_Template.xlsx"
    create_template(out_path)
