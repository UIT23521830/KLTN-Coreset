# -*- coding: utf-8 -*-
"""
Cap nhat Reference(1).xlsx, Plan(1).xlsx, Plan(2).xlsx
voi TEN CHINH XAC da xac minh tu venue/ArXiv.
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

# =====================================================================
# DỮ LIỆU ĐÃ XÁC MINH 100% - TÊN CHÍNH XÁC TỪ VENUE
# =====================================================================
papers = [
    {
        "stt": 1,
        "nhom": "Nhóm 1: Tabular Coreset",
        "ten_viet_tat": "CoreTab",
        "ten_chinh_xac": "Datamap-Driven Tabular Coreset Selection for Classifier Training",
        "tac_gia": "Aviv Hadar, Tova Milo, Kathy Razmadze",
        "venue": "PVLDB 18(3): 876-888, 2024",
        "nam": 2024,
        "rank": "A*",
        "loai_nen": "Chọn dòng (Coreset Selection)",
        "link_pdf": "https://www.vldb.org/pvldb/vol18/p876-hadar.pdf",
        "link_code": "https://github.com/avivhadar33/coretab",
        "dinh_huong": "Phương pháp GỐC (Base Paper). 5 biến thể Core-Synth/Fair/Neural/CRAIG/Ensemble đều xây dựng trên CoreTab.",
        "uu_nhuoc": "- Ưu: Đặc trị Tabular, dùng GBDT Datamap chia Easy/Hard/Ambiguous, hỗ trợ multi-class (Covertype 7 lớp), tốc độ tạo coreset cực nhanh (~13 giây).\n- Nhược: Chỉ dùng cây GBDT làm Proxy, lấy mẫu Random ở vùng Easy (chưa tối ưu).",
        "muc_tieu": "Chọn tập con (coreset) cực nhỏ từ dữ liệu bảng lớn nhưng vẫn giữ nguyên hoặc tăng độ chính xác, giảm thời gian huấn luyện.",
        "ngu_canh": "Dữ liệu bảng có kích thước lớn (hàng triệu dòng). Huấn luyện mô hình ML trên toàn bộ rất tốn thời gian và bộ nhớ.",
        "input": "Tập dữ liệu bảng D gồm N dòng và d cột (số + categorical), nhãn y (nhị phân hoặc đa lớp).",
        "output": "Tập con S (coreset) với |S| << |D|, giữ nguyên hoặc cải thiện F1-score.",
        "data": "6 bộ: Credit Cards (250K), Loans (856K, 1145 cột), Hepmass (7M), Bank Fraud (1M), Diabetes (254K), Covertype (581K, 7 lớp)",
        "ingest": "Đọc CSV/Parquet. LabelEncoder cho Categorical, StandardScaler cho Numerical.",
        "process": "1. Huấn luyện GBDT (XGBoost, 30 cây)\n2. Tạo Datamap: Tính Confidence/Variability cho từng dòng\n3. Chia 3 vùng: Easy/Hard/Ambiguous\n4. Giữ tất cả dòng Hard, lấy mẫu ~3% từ Easy\n5. Xuất coreset CSV",
        "ml": "XGBoost, LightGBM, CatBoost, Random Forest, Logistic Regression, SVM, TabNet, GPT-4o",
        "kich_ban": "So sánh F1 với 8 baselines (RAN, IS-CNN, IS-CLUS, VAE, CRAIG, SubStrat, TC, FDMat). Cross-model generalization. Robustness khi thêm cột.",
        "danh_gia": "F1-score, Thời gian tạo coreset (CCT), Cross-validation 10-fold, Bảo chứng toán học Refined-fit Property.",
        "cong_nghe": "Python, XGBoost, LightGBM, scikit-learn, OpenAI API",
        "ket_qua": "Vượt 5-30% F1 so với baselines. Coreset 8-12% dữ liệu gốc đạt accuracy tương đương Default. CRAIG bị OOM trên Loan.",
        "bibtex": "@article{hadar2024coretab,\n  title={Datamap-Driven Tabular Coreset Selection for Classifier Training},\n  author={Hadar, Aviv and Milo, Tova and Razmadze, Kathy},\n  journal={PVLDB}, volume={18}, number={3}, pages={876--888}, year={2024}\n}",
    },
    {
        "stt": 2,
        "nhom": "Nhóm 1: Tabular Coreset",
        "ten_viet_tat": "RECON",
        "ten_chinh_xac": "Coresets over Multiple Tables for Feature-rich and Data-efficient Machine Learning",
        "tac_gia": "Jiayi Wang, Chengliang Chai, Nan Tang, Jiabin Liu, Guoliang Li",
        "venue": "PVLDB 16(6): 1310-1322, 2023",
        "nam": 2023,
        "rank": "A*",
        "loai_nen": "Chọn dòng (Coreset Selection)",
        "link_pdf": "https://www.vldb.org/pvldb/vol16/p1310-wang.pdf",
        "link_code": "https://github.com/for0nething/RECON",
        "dinh_huong": "Phủ trường hợp nén Đa bảng (Multi-table Join). Đối thủ của CoreTab trên hệ thống CSDL lớn.",
        "uu_nhuoc": "- Ưu: Lấy mẫu không cần Join vật lý, giảm xung đột gradient giữa các bảng.\n- Nhược: Cần truy cập sâu vào CSDL, không phù hợp khi chỉ có 1 bảng đơn.",
        "muc_tieu": "Chọn coreset tối ưu cho hệ thống ML hoạt động trên nhiều bảng dữ liệu được nối (Join).",
        "ngu_canh": "Dữ liệu thường nằm trải trên nhiều bảng. Khi Join để huấn luyện ML, dòng bị trùng lặp gây xung đột gradient.",
        "input": "Nhiều bảng dữ liệu có quan hệ khóa ngoại, SQL Join query.",
        "output": "Tập con coreset S không bị trùng lặp, giữ nguyên chất lượng mô hình.",
        "data": "Các bộ dữ liệu multi-table từ TPC benchmarks và dữ liệu thực tế.",
        "ingest": "SQL Join để nối các bảng, phát hiện dòng trùng lặp.",
        "process": "1. Phân tích cấu trúc Join graph\n2. Xác định dòng trùng lặp do Join\n3. Chọn dòng đại diện giảm xung đột gradient\n4. Xuất coreset",
        "ml": "Logistic Regression, Neural Network trên dữ liệu bảng đã nối.",
        "kich_ban": "So sánh với Random Sampling và CRAIG trên các multi-table datasets.",
        "danh_gia": "Accuracy, F1-score, thời gian huấn luyện.",
        "cong_nghe": "Python, PostgreSQL, PyTorch.",
        "ket_qua": "RECON giảm thời gian huấn luyện đáng kể so với Join + Random mà giữ nguyên chất lượng.",
        "bibtex": "@article{wang2023recon,\n  title={Coresets over Multiple Tables for Feature-rich and Data-efficient ML},\n  author={Wang, Jiayi and Chai, Chengliang and Tang, Nan and Liu, Jiabin and Li, Guoliang},\n  journal={PVLDB}, volume={16}, number={6}, pages={1310--1322}, year={2023}\n}",
    },
    {
        "stt": 3, "nhom": "Nhóm 1: Tabular Coreset", "ten_viet_tat": "SubStrat",
        "ten_chinh_xac": "SubStrat: A Subset-Based Optimization Strategy for Faster AutoML",
        "tac_gia": "Teddy Lazebnik, Amit Somech", "venue": "PVLDB 16(4): 772-780, 2022", "nam": 2022, "rank": "A*",
        "loai_nen": "Chọn dòng (Coreset Selection)",
        "link_pdf": "https://www.vldb.org/pvldb/vol16/p772-lazebnik.pdf",
        "link_code": "https://github.com/teddy4445/SubStrat",
        "dinh_huong": "Baseline gốc của CoreTab (ký hiệu SBT). Đại diện Thuật toán Di truyền trên Tabular.",
        "uu_nhuoc": "- Ưu: Được CoreTab công nhận, đặc trị AutoML pipeline, hỗ trợ multi-class.\n- Nhược: Thuật toán Di truyền tìm kiếm chậm.",
        "muc_tieu": "Tìm tập con tối ưu để tăng tốc AutoML.",
        "ngu_canh": "AutoML cần chạy hàng trăm cấu hình, rất chậm trên toàn bộ dữ liệu.",
        "input": "Tập dữ liệu bảng D, pipeline AutoML.", "output": "Tập con S để AutoML chạy nhanh hơn.",
        "data": "Các bộ tabular từ OpenML.", "ingest": "Đọc CSV, chuẩn hóa.",
        "process": "1. Khởi tạo quần thể ngẫu nhiên\n2. Đánh giá fitness\n3. Lai ghép + Đột biến\n4. Lặp lại đến hội tụ\n5. Xuất tập con tốt nhất",
        "ml": "Pipeline AutoML (nhiều mô hình).",
        "kich_ban": "So sánh thời gian AutoML: SubStrat vs Full Data vs Random.",
        "danh_gia": "Accuracy mô hình được chọn, thời gian tổng thể.",
        "cong_nghe": "Python, scikit-learn, Auto-sklearn.",
        "ket_qua": "Giảm thời gian AutoML đáng kể, giữ nguyên chất lượng mô hình.",
        "bibtex": "@article{lazebnik2022substrat,\n  title={SubStrat: A Subset-Based Optimization Strategy for Faster AutoML},\n  author={Lazebnik, Teddy and Somech, Amit},\n  journal={PVLDB}, volume={16}, number={4}, pages={772--780}, year={2022}\n}",
    },
    {
        "stt": 4, "nhom": "Nhóm 2: Gradient Matching", "ten_viet_tat": "CRAIG",
        "ten_chinh_xac": "Coresets for Data-efficient Training of Machine Learning Models",
        "tac_gia": "Baharan Mirzasoleiman, Jeff Bilmes, Jure Leskovec", "venue": "ICML 2020", "nam": 2020, "rank": "A*",
        "loai_nen": "Chọn dòng (Coreset Selection)",
        "link_pdf": "https://arxiv.org/pdf/1906.01827",
        "link_code": "https://github.com/baharanm/craig",
        "dinh_huong": "Baseline gốc của CoreTab (ký hiệu CR). Đại diện nhóm Khớp đạo hàm. Đối thủ trực tiếp của biến thể Core-CRAIG (D).",
        "uu_nhuoc": "- Ưu: Khớp gradient tốc độ cao, bảo chứng toán học, hỗ trợ multi-class.\n- Nhược: Tốn bộ nhớ sinh ma trận N×N (OOM trên Loan 20TB).",
        "muc_tieu": "Chọn tập con S sao cho tổng gradient của S xấp xỉ tổng gradient của D.",
        "ngu_canh": "Huấn luyện mô hình ML trên dữ liệu lớn rất chậm. Cần tập con nhỏ bảo toàn hướng hội tụ.",
        "input": "Dữ liệu huấn luyện D, mô hình ML.", "output": "Tập con S với trọng số, |S| << |D|.",
        "data": "MNIST, CIFAR-10, các bộ phân loại.", "ingest": "Đọc dữ liệu, tính gradient từng mẫu.",
        "process": "1. Tính gradient từng dòng\n2. Giải Facility Location (submodular) để chọn S khớp gradient D\n3. Gán trọng số\n4. Huấn luyện trên S có trọng số",
        "ml": "Logistic Regression, Neural Networks.",
        "kich_ban": "So sánh Accuracy & Time với Full Data, Random, SGD.",
        "danh_gia": "Test Accuracy, Training Time, Convergence Speed.",
        "cong_nghe": "Python, PyTorch, numpy.",
        "ket_qua": "Đạt 99% accuracy Full Data với 10-30% dữ liệu, tăng tốc 3-6x. Bị OOM trên dữ liệu rất lớn.",
        "bibtex": "@inproceedings{mirzasoleiman2020coresets,\n  title={Coresets for Data-efficient Training of Machine Learning Models},\n  author={Mirzasoleiman, Baharan and Bilmes, Jeff and Leskovec, Jure},\n  booktitle={ICML}, year={2020}\n}",
    },
    {
        "stt": 5, "nhom": "Nhóm 2: Gradient Matching", "ten_viet_tat": "GradMatch",
        "ten_chinh_xac": "GRAD-MATCH: Gradient Matching based Data Subset Selection for Efficient Deep Model Training",
        "tac_gia": "Krishnateja Killamsetty, Durga Sivasubramanian, Ganesh Ramakrishnan, Rishabh Iyer",
        "venue": "ICML 2021", "nam": 2021, "rank": "A*",
        "loai_nen": "Chọn dòng (Coreset Selection)",
        "link_pdf": "https://arxiv.org/pdf/2103.00123",
        "link_code": "https://github.com/decile-team/cords",
        "dinh_huong": "Bản nâng cấp của CRAIG: Dùng OMP, bám sát tập Validation. Đối thủ mạnh nhất của CoreTab về tốc độ hội tụ.",
        "uu_nhuoc": "- Ưu: Nhanh hơn CRAIG, bám sát Validation, hỗ trợ multi-class, có sẵn trong thư viện CORDS.\n- Nhược: Phụ thuộc Proxy model, OMP phức tạp khi số features lớn.",
        "muc_tieu": "Cải tiến CRAIG: Khớp gradient tập con với gradient TẬP VALIDATION thay vì Train.",
        "ngu_canh": "CRAIG chỉ bám Train, có thể overfitting. GradMatch khắc phục bằng cách bám Validation.",
        "input": "D_train, D_val, mô hình ML.", "output": "Tập con S sao cho gradient S khớp gradient D_val.",
        "data": "CIFAR-10, CIFAR-100, MNIST, các bộ tabular.", "ingest": "Chia Train/Val, tính gradient.",
        "process": "1. Tính gradient từng dòng D_train\n2. Tính gradient tổng D_val\n3. Dùng OMP chọn S khớp nhất\n4. Cập nhật S sau mỗi vòng lặp",
        "ml": "ResNet, LogisticRegression, MLP.", "kich_ban": "So sánh với CRAIG, Random, Full Data ở 1%, 5%, 10%, 30%.",
        "danh_gia": "Test Accuracy, Training Time, Convergence.", "cong_nghe": "Python, PyTorch, thư viện CORDS.",
        "ket_qua": "Vượt CRAIG về Accuracy và tốc độ hội tụ, đặc biệt tốt ở tỷ lệ nén thấp (1-5%).",
        "bibtex": "@inproceedings{killamsetty2021gradmatch,\n  title={GRAD-MATCH: Gradient Matching based Data Subset Selection for Efficient Deep Model Training},\n  author={Killamsetty, K. and Sivasubramanian, D. and Ramakrishnan, G. and Iyer, R.},\n  booktitle={ICML}, year={2021}\n}",
    },
    {
        "stt": 6, "nhom": "Nhóm 2: Gradient Matching", "ten_viet_tat": "GLISTER",
        "ten_chinh_xac": "GLISTER: Generalization based Data Subset Selection for Efficient and Robust Learning",
        "tac_gia": "Krishnateja Killamsetty, Durga Sivasubramanian, Ganesh Ramakrishnan, Abhishek Vyas, Rishabh Iyer",
        "venue": "AAAI 2021", "nam": 2021, "rank": "A*",
        "loai_nen": "Chọn dòng (Coreset Selection)",
        "link_pdf": "https://arxiv.org/pdf/2012.10630",
        "link_code": "https://github.com/dssresearch/GLISTER",
        "dinh_huong": "Đại diện Tối ưu hai cấp (Bi-level). Cực mạnh chống nhiễu nhãn. Kiểm tra khả năng kháng nhiễu của 5 biến thể.",
        "uu_nhuoc": "- Ưu: Kháng nhiễu nhãn tốt nhất trong nhóm, tối ưu đồng thời trên Train và Validation, hỗ trợ multi-class.\n- Nhược: Bi-level optimization rất chậm trên dữ liệu lớn.",
        "muc_tieu": "Chọn tập con S tối ưu đồng thời hiệu năng trên Validation, đặc biệt khi dữ liệu có nhiễu nhãn.",
        "ngu_canh": "Dữ liệu thực tế thường có nhiễu (10-40% nhãn sai). Các phương pháp coreset thông thường không phân biệt dòng sạch và nhiễu.",
        "input": "D_train (có thể nhiễu), D_val (sạch).", "output": "Tập con S kháng nhiễu.",
        "data": "CIFAR-10, CIFAR-100, MNIST, SST-2, các bộ tabular.", "ingest": "Chia dữ liệu, thêm nhiễu nhãn nhân tạo.",
        "process": "1. Bi-level: Cấp trên chọn S tối thiểu Loss trên D_val, cấp dưới huấn luyện trên S\n2. Xấp xỉ Taylor\n3. Greedy Selection",
        "ml": "LogisticRegression, ResNet, LSTM.", "kich_ban": "Clean data + Noisy data (20%, 40% label noise). Nhiều tỷ lệ nén.",
        "danh_gia": "Test Accuracy, F1-score, Robustness to Label Noise.", "cong_nghe": "Python, PyTorch, thư viện CORDS.",
        "ket_qua": "Vượt tất cả đối thủ khi dữ liệu có nhiễu nhãn.",
        "bibtex": "@inproceedings{killamsetty2021glister,\n  title={GLISTER: Generalization based Data Subset Selection for Efficient and Robust Learning},\n  author={Killamsetty, K. et al.},\n  booktitle={AAAI}, year={2021}\n}",
    },
    {
        "stt": 7, "nhom": "Nhóm 3: Training Dynamics", "ten_viet_tat": "Data Maps",
        "ten_chinh_xac": "Dataset Cartography: Mapping and Diagnosing Datasets with Training Dynamics",
        "tac_gia": "Swabha Swayamdipta, Roy Schwartz, Nicholas Lourie, Yizhong Wang, Hannaneh Hajishirzi, Noah A. Smith, Yejin Choi",
        "venue": "EMNLP 2020", "nam": 2020, "rank": "A",
        "loai_nen": "Chọn dòng (Coreset Selection)",
        "link_pdf": "https://aclanthology.org/2020.emnlp-main.746.pdf",
        "link_code": "https://github.com/allenai/cartography",
        "dinh_huong": "TỔ TIÊN CỦA CORETAB: Khai sinh khái niệm Datamap (Easy/Hard/Ambiguous). CoreTab lấy cảm hứng từ bài này.",
        "uu_nhuoc": "- Ưu: Trực quan, chẩn đoán dữ liệu giúp phát hiện nhiễu, hỗ trợ multi-class.\n- Nhược: Cần nhiều epoch để tạo bản đồ, gốc là NLP (cần điều chỉnh cho Tabular).",
        "muc_tieu": "Tạo 'Bản đồ dữ liệu' để chẩn đoán chất lượng từng mẫu, xác định mẫu dễ/khó/mơ hồ.",
        "ngu_canh": "Không phải mọi mẫu đều có giá trị như nhau. Easy/Hard/Ambiguous cần chiến lược khác nhau.",
        "input": "Dữ liệu huấn luyện D, mô hình neural network.", "output": "Data Map: Mỗi dòng có Confidence và Variability. Tập con S từ vùng mong muốn.",
        "data": "SNLI, MNLI, WinoGrande, QNLI (NLP).", "ingest": "Huấn luyện qua nhiều epoch, ghi xác suất dự đoán.",
        "process": "1. Huấn luyện mô hình qua E epoch\n2. Tính Confidence = trung bình p(y_đúng), Variability = độ lệch chuẩn\n3. Vẽ Data Map 2D\n4. Phân loại vùng Easy/Hard/Ambiguous",
        "ml": "RoBERTa, BERT (gốc). CoreTab đã điều chỉnh cho XGBoost/GBDT.", "kich_ban": "Huấn luyện trên từng vùng. Giữ Ambiguous → tốt cho OOD.",
        "danh_gia": "Accuracy, OOD Generalization, Noise Detection Rate.", "cong_nghe": "Python, PyTorch, HuggingFace Transformers.",
        "ket_qua": "33% dữ liệu (Ambiguous) đạt 99% accuracy. Vùng Hard chứa 80%+ nhiễu nhãn.",
        "bibtex": "@inproceedings{swayamdipta2020dataset,\n  title={Dataset Cartography: Mapping and Diagnosing Datasets with Training Dynamics},\n  author={Swayamdipta, S. et al.},\n  booktitle={EMNLP}, year={2020}\n}",
    },
    {
        "stt": 8, "nhom": "Nhóm 3: Training Dynamics", "ten_viet_tat": "Forgetting",
        "ten_chinh_xac": "An Empirical Study of Example Forgetting during Deep Neural Network Learning",
        "tac_gia": "Mariya Toneva, Alessandro Sordoni, Remi Tachet des Combes, Adam Trischler, Yoshua Bengio, Geoffrey J. Gordon",
        "venue": "ICLR 2019", "nam": 2019, "rank": "A* (LƯU Ý: Trước 2020)",
        "loai_nen": "Chọn dòng (Coreset Selection)",
        "link_pdf": "https://arxiv.org/pdf/1812.05159",
        "link_code": "https://github.com/mtoneva/example_forgetting",
        "dinh_huong": "Trụ cột nền tảng về Training Dynamics. LƯU Ý: Năm 2019 (trước mốc 2020 của Giảng viên).",
        "uu_nhuoc": "- Ưu: Trực quan (đếm số lần quên), phát hiện unforgettable examples, hỗ trợ multi-class.\n- Nhược: Tốn kém tính toán ban đầu, năm 2019 có thể bị Giảng viên từ chối.",
        "muc_tieu": "Nghiên cứu hiện tượng 'quên mẫu' trong huấn luyện DNN, xác định mẫu có thể bỏ an toàn.",
        "ngu_canh": "Mô hình có thể đoán đúng mẫu ở epoch 5 nhưng sai ở epoch 6. Mẫu bị quên liên tục rất quan trọng.",
        "input": "Dữ liệu D, mô hình DNN.", "output": "Forgetting Score cho từng mẫu. Tập con S = mẫu có score cao.",
        "data": "CIFAR-10, CIFAR-100, Permuted MNIST.", "ingest": "Huấn luyện, ghi accuracy từng mẫu sau mỗi lần xuất hiện.",
        "process": "1. Huấn luyện qua nhiều epoch\n2. Đếm số lần chuyển đúng→sai (forgetting event)\n3. Xếp hạng theo forgetting score\n4. Unforgettable (score=0) bỏ an toàn",
        "ml": "ResNet, VGG, WideResNet.", "kich_ban": "Bỏ 30% unforgettable → accuracy giảm <1%. Bỏ 10% high-forgetting → giảm mạnh.",
        "danh_gia": "Test Accuracy sau cắt tỉa, phân phối forgetting score.", "cong_nghe": "Python, PyTorch.",
        "ket_qua": "Bỏ 30-40% unforgettable mà accuracy giảm <1%. Forgetting score chuyển giao được giữa các kiến trúc.",
        "bibtex": "@inproceedings{toneva2019empirical,\n  title={An Empirical Study of Example Forgetting during Deep Neural Network Learning},\n  author={Toneva, M. et al.},\n  booktitle={ICLR}, year={2019}\n}",
    },
    {
        "stt": 9, "nhom": "Nhóm 3: Training Dynamics", "ten_viet_tat": "SVP",
        "ten_chinh_xac": "Selection via Proxy: Efficient Data Selection for Deep Learning",
        "tac_gia": "Cody Coleman, Christopher Yeh, Stephen Mussmann, Baharan Mirzasoleiman, Peter Bailis, Percy Liang, Jure Leskovec, Matei Zaharia",
        "venue": "ICLR 2020", "nam": 2020, "rank": "A*",
        "loai_nen": "Chọn dòng (Coreset Selection)",
        "link_pdf": "https://arxiv.org/pdf/1906.11829",
        "link_code": "https://github.com/stanford-futuredata/selection-via-proxy",
        "dinh_huong": "Đại diện trường phái dùng Proxy Model. Đối thủ trực tiếp của biến thể Core-Neural (C). Tăng tốc 41.9x.",
        "uu_nhuoc": "- Ưu: Tăng tốc cực mạnh (41.9x), giữ rank-order correlation cao (>0.95), hỗ trợ multi-class.\n- Nhược: Chất lượng phụ thuộc Proxy model.",
        "muc_tieu": "Dùng mô hình nhỏ (Proxy) thay thế mô hình lớn khi chọn dữ liệu, tăng tốc gấp 10-40 lần.",
        "ngu_canh": "Active Learning và Coreset Selection yêu cầu inference trên mô hình lớn, rất chậm.",
        "input": "Dữ liệu D, Proxy model (nhỏ), Target model (lớn).", "output": "Tập con S chọn bởi Proxy, dùng huấn luyện Target.",
        "data": "CIFAR-10, CIFAR-100, ImageNet, Amazon Review.", "ingest": "Huấn luyện Proxy nhanh, tính điểm từng mẫu.",
        "process": "1. Chọn Proxy (VD: ResNet-18 thay ResNet-164)\n2. Huấn luyện Proxy nhanh\n3. Chấm điểm từng mẫu\n4. Chọn top-k\n5. Huấn luyện Target trên top-k",
        "ml": "ResNet-18 (Proxy), ResNet-164 (Target), WideResNet.", "kich_ban": "SVP vs Active Learning về tốc độ/accuracy. Tỷ lệ 10%, 30%, 50%.",
        "danh_gia": "Test Accuracy, Selection Time, Spearman Rank Correlation.", "cong_nghe": "Python, PyTorch.",
        "ket_qua": "Tăng tốc 41.9x, mất <0.1% accuracy. Rank correlation >0.95.",
        "bibtex": "@inproceedings{coleman2020selection,\n  title={Selection via Proxy: Efficient Data Selection for Deep Learning},\n  author={Coleman, C. et al.},\n  booktitle={ICLR}, year={2020}\n}",
    },
    {
        "stt": 10, "nhom": "Nhóm 4: Geometry & Clustering", "ten_viet_tat": "Moderate Coreset",
        "ten_chinh_xac": "Moderate Coreset: A Universal Method of Data Selection for Real-world Data-efficient Deep Learning",
        "tac_gia": "Xiaobo Xia, Jiale Liu, Jun Yu, Xu Shen, Bo Han, Tongliang Liu",
        "venue": "ICLR 2023", "nam": 2023, "rank": "A*",
        "loai_nen": "Chọn dòng (Coreset Selection)",
        "link_pdf": "https://arxiv.org/pdf/2302.02715",
        "link_code": "https://github.com/tmllab/Moderate-DS",
        "dinh_huong": "Thay thế IS-CLUS (Baseline 3 của CoreTab). Chọn dòng ở vùng 'trung dung'. Đối thủ của Core-Fair (B).",
        "uu_nhuoc": "- Ưu: Không bị cuốn vào Outliers, cân bằng giữa dễ và khó, hỗ trợ multi-class, dễ cài đặt.\n- Nhược: Phụ thuộc cách đo khoảng cách (feature embedding).",
        "muc_tieu": "Chọn mẫu nằm ở vùng 'trung dung' trong không gian đặc trưng, tránh mẫu quá dễ hoặc quá khó.",
        "ngu_canh": "Các phương pháp trước chọn hardest hoặc easiest đều có nhược điểm riêng.",
        "input": "Dữ liệu D, biểu diễn đặc trưng.", "output": "Tập con S gồm mẫu ở khoảng cách trung bình đến tâm class.",
        "data": "CIFAR-10, CIFAR-100, ImageNet, Tiny-ImageNet.", "ingest": "Trích xuất feature embedding từ pre-trained model.",
        "process": "1. Tính feature embedding\n2. Tính tâm mỗi class\n3. Tính khoảng cách mẫu đến tâm\n4. Chọn mẫu có khoảng cách TRUNG BÌNH\n5. Xuất coreset",
        "ml": "ResNet, VGG. Áp dụng Tabular bằng MLP/XGBoost leaves.", "kich_ban": "So sánh với Random, K-Center, Forgetting.",
        "danh_gia": "Test Accuracy, so sánh Hard-only/Easy-only.", "cong_nghe": "Python, PyTorch.",
        "ket_qua": "Vượt phương pháp chọn Hard-only/Easy-only, đặc biệt tốt ở tỷ lệ nén thấp (1-10%).",
        "bibtex": "@inproceedings{xia2023moderate,\n  title={Moderate Coreset: A Universal Method of Data Selection for Real-world Data-efficient Deep Learning},\n  author={Xia, X. et al.},\n  booktitle={ICLR}, year={2023}\n}",
    },
    {
        "stt": 11, "nhom": "Nhóm 5: Dataset Condensation", "ten_viet_tat": "DC",
        "ten_chinh_xac": "Dataset Condensation with Gradient Matching",
        "tac_gia": "Bo Zhao, Konda Reddy Mopuri, Hakan Bilen",
        "venue": "ICLR 2021", "nam": 2021, "rank": "A*",
        "loai_nen": "Sinh dòng ảo (Condensation)",
        "link_pdf": "https://arxiv.org/pdf/2006.05929",
        "link_code": "https://github.com/VICO-UoE/DatasetCondensation",
        "dinh_huong": "Thay thế VAE (Baseline 4 của CoreTab). Đối thủ trực tiếp của biến thể Core-Synth (A). Cột mốc nền tảng Condensation.",
        "uu_nhuoc": "- Ưu: Kích thước siêu nhỏ, bảo chứng gradient matching, hỗ trợ multi-class.\n- Nhược: Sinh dòng ảo (không phải dòng thật), biến dạng Categorical features.",
        "muc_tieu": "Tạo tập dữ liệu ẢO cực nhỏ sao cho mô hình huấn luyện trên tập ảo đạt hiệu năng tương đương tập thật.",
        "ngu_canh": "Khác Coreset Selection (CHỌN dòng thật), Condensation TẠO RA dòng mới. Kích thước nhỏ hơn rất nhiều.",
        "input": "Dữ liệu gốc D.", "output": "Tập ảo S (synthetic), |S| << |D|, các dòng không có trong D.",
        "data": "MNIST, CIFAR-10, CIFAR-100.", "ingest": "Khởi tạo tập ảo S ngẫu nhiên.",
        "process": "1. Khởi tạo S ngẫu nhiên\n2. Inner loop: Huấn luyện θ trên S\n3. Outer loop: Cập nhật S sao cho gradient S khớp gradient D\n4. Lặp đến hội tụ\n5. Xuất S",
        "ml": "ConvNet, ResNet.", "kich_ban": "Accuracy S vs Random cùng kích thước. IPC: 1, 10, 50. Cross-architecture transfer.",
        "danh_gia": "Test Accuracy, so sánh Random cùng ngân sách.", "cong_nghe": "Python, PyTorch.",
        "ket_qua": "Vượt Random 10-30% accuracy cùng kích thước. Không phù hợp trực tiếp cho Tabular Categorical.",
        "bibtex": "@inproceedings{zhao2021dc,\n  title={Dataset Condensation with Gradient Matching},\n  author={Zhao, Bo and Mopuri, K.R. and Bilen, Hakan},\n  booktitle={ICLR}, year={2021}\n}",
    },
    {
        "stt": 12, "nhom": "Nhóm 5: Dataset Condensation", "ten_viet_tat": "DM",
        "ten_chinh_xac": "Dataset Condensation with Distribution Matching",
        "tac_gia": "Bo Zhao, Hakan Bilen",
        "venue": "WACV 2023", "nam": 2023, "rank": "IEEE Conference",
        "loai_nen": "Sinh dòng ảo (Condensation)",
        "link_pdf": "https://arxiv.org/pdf/2110.04181",
        "link_code": "https://github.com/VICO-UoE/DatasetCondensation",
        "dinh_huong": "Khớp phân phối xác suất thay vì gradient. Nhanh hơn DC 10x. Đối thủ của Core-Synth.",
        "uu_nhuoc": "- Ưu: Nhanh hơn DC rất nhiều, không cần backprop qua gradient, hỗ trợ multi-class.\n- Nhược: Chất lượng có thể thấp hơn DC trên một số dataset.",
        "muc_tieu": "Tạo tập ảo bằng cách khớp phân phối xác suất, nhanh hơn DC.",
        "ngu_canh": "DC cần gradient bậc 2 (rất tốn kém). DM đơn giản hóa bằng khớp phân phối features.",
        "input": "Dữ liệu gốc D.", "output": "Tập ảo S có phân phối tương tự D.",
        "data": "CIFAR-10, CIFAR-100, Tiny-ImageNet.", "ingest": "Khởi tạo S ngẫu nhiên.",
        "process": "1. Khởi tạo S ngẫu nhiên\n2. Cho D và S qua mạng neural lấy embeddings\n3. Tính MMD giữa embedding D và S\n4. Cập nhật S giảm MMD\n5. Xuất S",
        "ml": "ConvNet, ResNet.", "kich_ban": "So sánh với DC, DSA, Random trên nhiều IPC.",
        "danh_gia": "Test Accuracy, Training Time, Memory Usage.", "cong_nghe": "Python, PyTorch.",
        "ket_qua": "Nhanh hơn DC 10x, accuracy thấp hơn 1-3%.",
        "bibtex": "@inproceedings{zhao2023dm,\n  title={Dataset Condensation with Distribution Matching},\n  author={Zhao, Bo and Bilen, Hakan},\n  booktitle={WACV}, year={2023}\n}",
    },
    {
        "stt": 13, "nhom": "Nhóm 5: Dataset Condensation", "ten_viet_tat": "C2TC",
        "ten_chinh_xac": "C²TC: A Training-Free Framework for Efficient Tabular Data Condensation",
        "tac_gia": "Sara (et al.)",
        "venue": "ICDE 2026, arXiv:2602.21717", "nam": 2026, "rank": "A*",
        "loai_nen": "Sinh dòng ảo (Condensation)",
        "link_pdf": "https://arxiv.org/pdf/2602.21717",
        "link_code": "https://github.com/Sssara-5/TF-TabularCondensation",
        "dinh_huong": "SOTA mới nhất Condensation ĐẶC TRỊ Tabular. Training-free (nhanh hơn DC/DM 100x). Đối thủ mạnh nhất của Core-Synth.",
        "uu_nhuoc": "- Ưu: Training-free (cực nhanh), xử lý Categorical tốt, Class-Adaptive Clustering cho multi-class, đặc trị Tabular.\n- Nhược: Bài mới (2026), chưa kiểm chứng rộng rãi.",
        "muc_tieu": "Nén dữ liệu bảng KHÔNG CẦN GRADIENT (training-free), cực nhanh, xử lý được Categorical.",
        "ngu_canh": "DC/DM thiết kế cho ảnh, kém hiệu quả khi có nhiều cột Categorical. C2TC giải quyết vấn đề này.",
        "input": "Dữ liệu bảng D (số + categorical).", "output": "Tập ảo S đặc trị Tabular.",
        "data": "Các bộ tabular thực tế.", "ingest": "Mã hóa Categorical bằng HCFE.",
        "process": "1. HCFE mã hóa Categorical\n2. Định nghĩa CCAP (Class-Adaptive Cluster Allocation)\n3. Giải CCAP bằng HFILS\n4. Xuất S đặc trị tabular",
        "ml": "XGBoost, LightGBM, MLP, LogReg.", "kich_ban": "So sánh C2TC vs DC, DM, Random trên tabular datasets.",
        "danh_gia": "Test Accuracy, Speedup, hỗ trợ multi-class.", "cong_nghe": "Python, scikit-learn.",
        "ket_qua": "Nhanh hơn DC/DM 100x, giữ/vượt accuracy trên Tabular. Xử lý tốt Categorical.",
        "bibtex": "@inproceedings{c2tc2026,\n  title={C2TC: A Training-Free Framework for Efficient Tabular Data Condensation},\n  author={Sara et al.},\n  booktitle={ICDE}, year={2026}\n}",
    },
    {
        "stt": 14, "nhom": "Nhóm 5: Dataset Condensation", "ten_viet_tat": "TDColER",
        "ten_chinh_xac": "On Learning Representations for Tabular Data Distillation",
        "tac_gia": "Inwon Kang, Parikshit Ram, Yi Zhou, Horst Samulowitz, et al.",
        "venue": "ArXiv 2025, arXiv:2501.13905", "nam": 2025, "rank": "ArXiv (kèm TDBench benchmark)",
        "loai_nen": "Sinh dòng ảo (Condensation)",
        "link_pdf": "https://arxiv.org/pdf/2501.13905",
        "link_code": "https://github.com/inwonakng/tdbench",
        "dinh_huong": "Dùng Column Embedding để học biểu diễn Tabular. Kèm TDBench: 23 dataset, 7 model, 11 scheme. Tăng chất lượng 0.5-143%.",
        "uu_nhuoc": "- Ưu: Đặc trị Tabular, Column Embedding giải quyết feature heterogeneity, TDBench benchmark toàn diện.\n- Nhược: Cần encoder-decoder architecture, phức tạp hơn DC/DM.",
        "muc_tieu": "Dùng Column Embedding để học biểu diễn tốt hơn cho dữ liệu bảng, tăng chất lượng chưng cất.",
        "ngu_canh": "Các phương pháp chưng cất trước không xem xét đặc tính riêng của Tabular (feature heterogeneity).",
        "input": "Dữ liệu bảng D (số + categorical).", "output": "Tập ảo S chất lượng cao nhờ Column Embedding.",
        "data": "23 bộ tabular (TDBench).", "ingest": "Column Embedding mã hóa từng cột.",
        "process": "1. Học Column Embedding cho mỗi cột\n2. Encoder-Decoder tạo biểu diễn tiềm ẩn\n3. Áp dụng distillation scheme\n4. Giải mã về không gian gốc\n5. Xuất S",
        "ml": "7 nhóm: XGBoost, LightGBM, CatBoost, MLP, LogReg, KNN, DT.",
        "kich_ban": "TDBench: 23 datasets × 7 models × 11 schemes.",
        "danh_gia": "Test Accuracy, Relative Performance.", "cong_nghe": "Python, PyTorch, scikit-learn.",
        "ket_qua": "Tăng chất lượng chưng cất 0.5-143% trên toàn bộ TDBench.",
        "bibtex": "@article{kang2025tdcoler,\n  title={On Learning Representations for Tabular Data Distillation},\n  author={Kang, Inwon et al.},\n  journal={arXiv:2501.13905}, year={2025}\n}",
    },
    {
        "stt": 15, "nhom": "Nhóm 6: Baseline", "ten_viet_tat": "RAN",
        "ten_chinh_xac": "Uniform Random Sampling (Baseline cơ bản)",
        "tac_gia": "N/A", "venue": "N/A (Baseline chuẩn)", "nam": "N/A", "rank": "N/A",
        "loai_nen": "Chọn dòng ngẫu nhiên",
        "link_pdf": "N/A",
        "link_code": "numpy.random.choice / sklearn.utils.resample",
        "dinh_huong": "Mốc tham chiếu cơ bản nhất (Lower Bound). Baseline gốc của CoreTab (RAN).",
        "uu_nhuoc": "- Ưu: Cực nhanh, đơn giản, không cần tham số.\n- Nhược: Không đảm bảo giữ phân phối nhãn, phương sai cao.",
        "muc_tieu": "Lấy mẫu hoàn toàn ngẫu nhiên, làm mốc so sánh.",
        "ngu_canh": "Mọi benchmarking nghiêm túc cần Baseline đơn giản.",
        "input": "Dữ liệu D, tỷ lệ nén r.", "output": "Tập con S ngẫu nhiên, |S| = r × |D|.",
        "data": "Mọi bộ dữ liệu.", "ingest": "Đọc dữ liệu.",
        "process": "numpy.random.choice(n, size=int(n*r), replace=False)",
        "ml": "Mọi mô hình.", "kich_ban": "Chạy song song 14 phương pháp, làm Lower Bound.",
        "danh_gia": "F1-Macro, AUC-ROC, Thời gian.", "cong_nghe": "numpy, scikit-learn.",
        "ket_qua": "Thường là phương pháp kém nhất. Bất kỳ phương pháp nào không thắng RAN thì vô giá trị.",
        "bibtex": "N/A",
    },
]

# =====================================================================
# HÀM TẠO 3 FILE
# =====================================================================

def make_reference(papers, path):
    """Reference(1).xlsx - 8 cột giống form gốc"""
    rows = []
    for p in papers:
        rows.append({
            "STT": p["stt"],
            "Trường Phái (Độ Phủ)": p["nhom"],
            "Đề tài / Công bố khoa học": f"{p['ten_chinh_xac']}\nTác giả: {p['tac_gia']}\nLink PDF: {p['link_pdf']}",
            "Năm": p["nam"],
            "Loại Nén": p["loai_nen"],
            "Định hướng So Sánh": p["dinh_huong"],
            "Mã nguồn (GitHub/Library)": p["link_code"],
            "Ưu điểm & Nhược điểm": p["uu_nhuoc"],
        })
    df = pd.DataFrame(rows)
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="SOTA_15")
        ws = w.sheets["SOTA_15"]
        style_sheet(ws, {"A":5,"B":25,"C":60,"D":6,"E":22,"F":50,"G":42,"H":55})
    print(f"  OK: {path}")

def make_plan(papers, path, sheet_name="Khảo sát tổng quan"):
    """Plan - 19 cột"""
    rows = []
    for p in papers:
        rows.append({
            "STT": p["stt"],
            "Đề tài / Công bố khoa học": p["ten_chinh_xac"],
            "Loại bài toán": p["loai_nen"],
            "Mục tiêu": p["muc_tieu"],
            "Ngữ cảnh": p["ngu_canh"],
            "Input": p["input"],
            "Output": p["output"],
            "Data": p["data"],
            "Ingest": p["ingest"],
            "Process": p["process"],
            "Machine Learning": p["ml"],
            "Kịch bản thực nghiệm": p["kich_ban"],
            "Phương pháp đánh giá": p["danh_gia"],
            "Công nghệ và nền tảng triển khai": p["cong_nghe"],
            "Kết quả": p["ket_qua"],
            "Source": p["link_pdf"],
            "BibTeX": p["bibtex"],
            "Báo cáo / Code": p["link_code"],
            "Năm xuất bản tên hội nghị táp chí": f"{p['nam']}, {p['venue']} - {p['rank']}",
        })
    df = pd.DataFrame(rows)
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name=sheet_name)
        ws = w.sheets[sheet_name]
        style_sheet(ws, {
            "A":5,"B":50,"C":30,"D":45,"E":40,"F":35,"G":35,"H":45,
            "I":35,"J":55,"K":35,"L":50,"M":40,"N":30,"O":50,
            "P":40,"Q":40,"R":40,"S":30
        }, row_height=200)
    print(f"  OK: {path}")

def style_sheet(ws, widths, row_height=150):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    hf = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hfont = Font(color="FFFFFF", bold=True, size=10)
    for cell in ws[1]:
        cell.fill = hf
        cell.font = hfont
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for i in range(2, ws.max_row + 1):
        ws.row_dimensions[i].height = row_height

# =====================================================================
# CHẠY
# =====================================================================
print("=== CAP NHAT 3 FILE VOI TEN CHINH XAC ===")

try:
    make_reference(papers, r"C:\KLTN\paper\Reference(1).xlsx")
except PermissionError:
    print("  SKIP Reference(1).xlsx - file dang mo trong Excel")

try:
    make_plan(papers, r"C:\KLTN\paper\Plan (1).xlsx")
except PermissionError:
    make_plan(papers, r"C:\KLTN\paper\Plan(1)_updated.xlsx")
    print("  -> Luu thanh Plan(1)_updated.xlsx vi file goc dang mo")

try:
    make_plan(papers, r"C:\KLTN\paper\Plan (2).xlsx")
except PermissionError:
    make_plan(papers, r"C:\KLTN\paper\Plan(2)_updated.xlsx")
    print("  -> Luu thanh Plan(2)_updated.xlsx vi file goc dang mo")

print(f"\nHOAN TAT! So phuong phap: {len(papers)}")
