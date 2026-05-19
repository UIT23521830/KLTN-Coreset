# -*- coding: utf-8 -*-
"""
Tạo Plan(2).xlsx = bản tiếng Việt CÓ DẤU của Plan(1).xlsx.
"""
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

data = [
    # ===== 1. CoreTab =====
    {
        "STT": 1,
        "Đề tài / Công bố khoa học": "CoreTab: Data-efficient Machine Learning over Tabular Data via Datamap-based Coreset Selection",
        "Loại bài toán": "Lựa chọn tập cốt lõi cho dữ liệu bảng (Coreset Selection for Tabular Data)",
        "Mục tiêu": "Chọn tập con (coreset) cực nhỏ từ dữ liệu bảng lớn nhưng vẫn giữ nguyên hoặc tăng độ chính xác của mô hình ML, giảm thời gian huấn luyện từ hàng giờ xuống hàng phút.",
        "Ngữ cảnh": "Dữ liệu bảng (tabular) có kích thước lớn (hàng triệu dòng), nhiều cột số và phân loại. Việc huấn luyện mô hình ML trên toàn bộ dữ liệu rất tốn thời gian và bộ nhớ.",
        "Input": "Tập dữ liệu bảng D gồm N dòng và d cột (số + categorical), nhãn y (nhị phân hoặc đa lớp).",
        "Output": "Tập con S (coreset) với |S| << |D|, giữ nguyên hoặc cải thiện F1-score khi huấn luyện mô hình trên S.",
        "Data": "6 bộ dữ liệu thực tế:\n1. Credit Cards (CC): 250K dòng, 0.17% nhãn dương\n2. Loans (LN): 856K dòng, 1145 cột\n3. Hepmass (HP): 7 triệu dòng\n4. Bank Fraud (BF): 1 triệu dòng\n5. Diabetes (DI): 254K dòng\n6. Covertype (CT): 581K dòng, 7 lớp (đa lớp)",
        "Ingest": "Đọc dữ liệu bảng từ file CSV/Parquet. Tiền xử lý: LabelEncoder cho Categorical, StandardScaler cho Numerical.",
        "Process": "1. Huấn luyện GBDT (XGBoost, 30 cây) trên dữ liệu.\n2. Tạo Datamap: Tính Confidence và Variability cho từng dòng qua các cây.\n3. Chia dữ liệu thành 3 vùng: Easy (độ chính xác cao), Hard (gần biên quyết định), Ambiguous (nhỏ, nhất quán).\n4. Giữ tất cả dòng Hard, lấy mẫu nhỏ (~3%) từ vùng Easy.\n5. Xuất coreset ra file CSV.",
        "Machine Learning": "XGBoost, LightGBM, CatBoost, Random Forest, Logistic Regression, SVM (RBF), TabNet, GPT-4o (fine-tune).",
        "Kịch bản thực nghiệm": "1. So sánh F1-score giữa CoreTab và 8 baselines (RAN, IS-CNN, IS-CLUS, VAE, CRAIG, SubStrat, TC, FDMat) tại các mức nén.\n2. Cross-model generalization: Nén bằng XGBoost, đánh giá bằng LightGBM/TabNet.\n3. Robustness: Thêm cột mới sau khi nén, kiểm tra coreset vẫn hiệu quả.\n4. Timeout 24h, ghi nhận OOM.",
        "Phương pháp đánh giá": "1. F1-score (Chỉ số chính)\n2. Thời gian tạo coreset (CCT - Coreset Creation Time)\n3. Cross-validation 10-fold\n4. Bảo chứng toán học: Refined-fit Property (giới hạn sai lệch Recall/Precision).",
        "Công nghệ và nền tảng triển khai": "Python, XGBoost, LightGBM, scikit-learn, OpenAI API (GPT-4o fine-tuning).",
        "Kết quả": "1. CoreTab vượt 5-30% F1-score so với baselines trên hầu hết datasets.\n2. Coreset chỉ chiếm 8-12% dữ liệu gốc nhưng đạt độ chính xác tương đương hoặc cao hơn Default (huấn luyện toàn bộ).\n3. CRAIG bị OOM trên Loan dataset (cần ma trận 20TB).\n4. Thời gian tạo coreset: 13 giây (CC), vài phút (các bộ khác).",
        "Source": "https://vldb.org/pvldb/vol18/p451-hadar.pdf",
        "BibTeX": "@article{hadar2024coretab,\n  title={CoreTab},\n  author={Hadar, Aviv and Milo, Tova and Razmadze, Kathy},\n  journal={PVLDB},\n  volume={18}, number={3},\n  year={2024}\n}",
        "Báo cáo / Code": "https://github.com/avivhadar33/coretab",
        "Năm xuất bản tên hội nghị táp chí": "2024, PVLDB (Proceedings of the VLDB Endowment) - Xếp hạng A*",
    },

    # ===== 2. RECON =====
    {
        "STT": 2,
        "Đề tài / Công bố khoa học": "RECON: Reducing Conflicting Gradients From Data for ML over Multi-Table Joins",
        "Loại bài toán": "Lựa chọn tập cốt lõi cho dữ liệu đa bảng (Coreset Selection for Multi-Table)",
        "Mục tiêu": "Chọn coreset tối ưu cho hệ thống ML hoạt động trên nhiều bảng dữ liệu được nối (Join), giảm xung đột gradient giữa các dòng trùng lặp.",
        "Ngữ cảnh": "Trong hệ thống CSDL thực tế, dữ liệu thường nằm trải trên nhiều bảng. Khi Join các bảng để huấn luyện ML, dòng bị trùng lặp gây ra xung đột gradient làm giảm chất lượng mô hình.",
        "Input": "Nhiều bảng dữ liệu có quan hệ khóa ngoại (Foreign Key), SQL Join query.",
        "Output": "Tập con coreset S không bị trùng lặp, giữ nguyên chất lượng mô hình.",
        "Data": "Các bộ dữ liệu multi-table từ TPC benchmarks và dữ liệu thực tế.",
        "Ingest": "SQL Join để nối các bảng, phát hiện dòng trùng lặp.",
        "Process": "1. Phân tích cấu trúc Join graph.\n2. Xác định các dòng bị trùng lặp do Join.\n3. Chọn dòng đại diện giảm xung đột gradient.\n4. Xuất coreset.",
        "Machine Learning": "Logistic Regression, Neural Network trên dữ liệu bảng đã nối.",
        "Kịch bản thực nghiệm": "So sánh với Random Sampling và CRAIG trên các multi-table datasets.",
        "Phương pháp đánh giá": "Accuracy, F1-score, thời gian huấn luyện.",
        "Công nghệ và nền tảng triển khai": "Python, PostgreSQL, PyTorch.",
        "Kết quả": "RECON giảm thời gian huấn luyện đáng kể so với Join + Random Sampling mà giữ nguyên chất lượng.",
        "Source": "https://vldb.org/pvldb/vol17/p3370-wang.pdf",
        "BibTeX": "@article{wang2024recon,\n  title={RECON},\n  author={Wang, Jiayi et al.},\n  journal={PVLDB},\n  volume={17}, number={11},\n  year={2024}\n}",
        "Báo cáo / Code": "https://github.com/for0nething/RECON",
        "Năm xuất bản tên hội nghị táp chí": "2024, PVLDB - Xếp hạng A*",
    },

    # ===== 3. SubStrat =====
    {
        "STT": 3,
        "Đề tài / Công bố khoa học": "SubStrat: A Subset-Based Optimization Strategy for Faster AutoML",
        "Loại bài toán": "Lựa chọn tập cốt lõi cho quy trình AutoML",
        "Mục tiêu": "Tìm tập con dữ liệu tối ưu để tăng tốc quá trình AutoML (tự động chọn mô hình + siêu tham số), giảm thời gian từ hàng giờ xuống hàng phút.",
        "Ngữ cảnh": "AutoML cần chạy hàng trăm cấu hình mô hình. Nếu mỗi cấu hình đều chạy trên toàn bộ dữ liệu thì cực chậm.",
        "Input": "Tập dữ liệu bảng D, pipeline AutoML.",
        "Output": "Tập con S để AutoML chạy nhanh hơn mà vẫn chọn đúng mô hình tốt nhất.",
        "Data": "Các bộ dữ liệu tabular từ OpenML.",
        "Ingest": "Đọc dữ liệu bảng từ CSV, chuẩn hóa.",
        "Process": "1. Khởi tạo quần thể (Population) ngẫu nhiên.\n2. Đánh giá fitness của từng tập con bằng độ chính xác mô hình.\n3. Lai ghép (Crossover) + Đột biến (Mutation) để tạo thế hệ mới.\n4. Lặp lại cho đến khi hội tụ.\n5. Xuất tập con tốt nhất.",
        "Machine Learning": "Pipeline AutoML (nhiều mô hình khác nhau).",
        "Kịch bản thực nghiệm": "So sánh thời gian AutoML khi dùng SubStrat vs Toàn bộ dữ liệu vs Lấy mẫu ngẫu nhiên.",
        "Phương pháp đánh giá": "Accuracy của mô hình được AutoML chọn, thời gian tổng thể.",
        "Công nghệ và nền tảng triển khai": "Python, scikit-learn, Auto-sklearn.",
        "Kết quả": "SubStrat giảm thời gian AutoML đáng kể mà giữ nguyên chất lượng mô hình được chọn. Được CoreTab công nhận là baseline hiệu quả cho Tabular.",
        "Source": "https://vldb.org/pvldb/vol16/p772-lazebnik.pdf",
        "BibTeX": "@article{lazebnik2022substrat,\n  title={SubStrat},\n  author={Lazebnik, Teddy and Somech, Amit},\n  journal={PVLDB},\n  volume={16}, number={4},\n  year={2022}\n}",
        "Báo cáo / Code": "https://github.com/teddy4445/SubStrat",
        "Năm xuất bản tên hội nghị táp chí": "2022, PVLDB - Xếp hạng A*",
    },

    # ===== 4. CRAIG =====
    {
        "STT": 4,
        "Đề tài / Công bố khoa học": "Coresets for Data-efficient Training of Machine Learning Models",
        "Loại bài toán": "Lựa chọn tập cốt lõi bằng Khớp đạo hàm (Gradient Matching)",
        "Mục tiêu": "Chọn tập con S sao cho tổng gradient của S xấp xỉ tổng gradient của toàn bộ D, đảm bảo mô hình hội tụ tương tự.",
        "Ngữ cảnh": "Huấn luyện mô hình ML (đặc biệt Logistic Regression và Neural Network) trên dữ liệu lớn rất chậm. Cần một tập con nhỏ mà vẫn bảo toàn hướng hội tụ.",
        "Input": "Dữ liệu huấn luyện D, mô hình ML cần huấn luyện.",
        "Output": "Tập con S với trọng số (weighted subset), |S| << |D|.",
        "Data": "MNIST, CIFAR-10, các bộ dữ liệu phân loại.",
        "Ingest": "Đọc dữ liệu, tính gradient của từng mẫu.",
        "Process": "1. Tính gradient của từng dòng dữ liệu với mô hình hiện tại.\n2. Giải bài toán Facility Location (hàm lồi/submodular) để chọn S sao cho gradient của S khớp gradient của D.\n3. Gán trọng số cho từng dòng trong S.\n4. Huấn luyện mô hình trên S (có trọng số).",
        "Machine Learning": "Logistic Regression (tối ưu), Neural Networks.",
        "Kịch bản thực nghiệm": "So sánh Accuracy và Training Time với Toàn bộ dữ liệu, Lấy mẫu ngẫu nhiên, SGD thông thường.",
        "Phương pháp đánh giá": "Test Accuracy, Thời gian huấn luyện, Tốc độ hội tụ.",
        "Công nghệ và nền tảng triển khai": "Python, PyTorch, numpy.",
        "Kết quả": "CRAIG đạt 99% accuracy của toàn bộ dữ liệu với chỉ 10-30% dữ liệu, tăng tốc 3-6 lần. Tuy nhiên bị OOM trên dữ liệu rất lớn (cần ma trận N×N).",
        "Source": "https://arxiv.org/abs/1906.01827",
        "BibTeX": "@inproceedings{mirzasoleiman2020coresets,\n  title={Coresets for Data-efficient Training},\n  author={Mirzasoleiman, B. and Bilmes, J. and Leskovec, J.},\n  booktitle={ICML},\n  year={2020}\n}",
        "Báo cáo / Code": "https://github.com/baharanm/craig",
        "Năm xuất bản tên hội nghị táp chí": "2020, ICML - Xếp hạng A*",
    },

    # ===== 5. GradMatch =====
    {
        "STT": 5,
        "Đề tài / Công bố khoa học": "GRAD-MATCH: Gradient Matching based Data Subset Selection for Efficient Deep Model Training",
        "Loại bài toán": "Lựa chọn tập cốt lõi bằng Khớp đạo hàm nâng cao",
        "Mục tiêu": "Cải tiến CRAIG: Khớp gradient của tập con với gradient của TẬP VALIDATION (thay vì tập Train) để tránh overfitting.",
        "Ngữ cảnh": "CRAIG chỉ bám sát gradient của tập Train, có thể dẫn đến overfitting. GradMatch khắc phục bằng cách bám sát Validation.",
        "Input": "Dữ liệu Train D_train, dữ liệu Validation D_val, mô hình ML.",
        "Output": "Tập con S của D_train sao cho gradient của S khớp gradient của D_val.",
        "Data": "CIFAR-10, CIFAR-100, MNIST, các bộ tabular.",
        "Ingest": "Chia dữ liệu Train/Val, tính gradient.",
        "Process": "1. Tính gradient của từng dòng trong D_train.\n2. Tính gradient tổng của D_val.\n3. Dùng Orthogonal Matching Pursuit (OMP) để chọn S sao cho gradient của S khớp nhất với gradient của D_val.\n4. Cập nhật S sau mỗi vòng lặp huấn luyện.",
        "Machine Learning": "ResNet, LogisticRegression, MLP, hỗ trợ mọi mô hình có gradient.",
        "Kịch bản thực nghiệm": "So sánh với CRAIG, Random, Toàn bộ dữ liệu trên nhiều dataset và nhiều tỷ lệ nén (1%, 5%, 10%, 30%).",
        "Phương pháp đánh giá": "Test Accuracy, Thời gian huấn luyện, Tốc độ hội tụ.",
        "Công nghệ và nền tảng triển khai": "Python, PyTorch, thư viện CORDS.",
        "Kết quả": "GradMatch vượt CRAIG về Accuracy và tốc độ hội tụ, đặc biệt tốt ở tỷ lệ nén thấp (1-5%).",
        "Source": "https://arxiv.org/abs/2103.00123",
        "BibTeX": "@inproceedings{killamsetty2021gradmatch,\n  title={GRAD-MATCH},\n  author={Killamsetty, K. et al.},\n  booktitle={ICML},\n  year={2021}\n}",
        "Báo cáo / Code": "https://github.com/decile-team/cords",
        "Năm xuất bản tên hội nghị táp chí": "2021, ICML - Xếp hạng A*",
    },

    # ===== 6. GLISTER =====
    {
        "STT": 6,
        "Đề tài / Công bố khoa học": "GLISTER: Generalization based Data Subset Selection for Efficient and Robust Learning",
        "Loại bài toán": "Lựa chọn tập cốt lõi bằng Tối ưu hai cấp (Bi-level Optimization)",
        "Mục tiêu": "Chọn tập con S tối ưu hóa ĐỒNG THỜI hiệu năng trên tập Validation, đặc biệt hiệu quả khi dữ liệu có nhiều nhãn sai (Label Noise).",
        "Ngữ cảnh": "Dữ liệu thực tế thường có nhiễu (10-40% nhãn sai). Các phương pháp coreset thông thường không phân biệt được dòng sạch và dòng nhiễu.",
        "Input": "Dữ liệu Train D_train (có thể có nhiễu), dữ liệu Validation D_val (sạch).",
        "Output": "Tập con S kháng nhiễu, tối ưu cho Generalization.",
        "Data": "CIFAR-10, CIFAR-100, MNIST, SST-2 (NLP), các bộ tabular.",
        "Ingest": "Chia dữ liệu, thêm nhiễu nhãn nhân tạo để kiểm tra.",
        "Process": "1. Định nghĩa bài toán Tối ưu hai cấp (Bi-level):\n   - Cấp trên: Chọn S tối thiểu Loss trên D_val.\n   - Cấp dưới: Huấn luyện mô hình trên S.\n2. Giải xấp xỉ bằng khai triển Taylor.\n3. Dùng Greedy Selection để chọn từng dòng vào S.",
        "Machine Learning": "LogisticRegression, ResNet, LSTM.",
        "Kịch bản thực nghiệm": "1. Dữ liệu sạch: So sánh Accuracy với CRAIG, GradMatch, Random.\n2. Dữ liệu nhiễu (20%, 40% nhãn sai): Kiểm tra độ kháng nhiễu.\n3. Nhiều tỷ lệ nén khác nhau.",
        "Phương pháp đánh giá": "Test Accuracy, F1-score, Độ kháng nhiễu nhãn (Robustness to Label Noise).",
        "Công nghệ và nền tảng triển khai": "Python, PyTorch, thư viện CORDS.",
        "Kết quả": "GLISTER vượt tất cả đối thủ khi dữ liệu có nhiễu nhãn. Trên dữ liệu sạch, tương đương hoặc tốt hơn GradMatch.",
        "Source": "https://arxiv.org/abs/2012.10630",
        "BibTeX": "@inproceedings{killamsetty2021glister,\n  title={GLISTER},\n  author={Killamsetty, K. et al.},\n  booktitle={AAAI},\n  year={2021}\n}",
        "Báo cáo / Code": "https://github.com/dssresearch/GLISTER",
        "Năm xuất bản tên hội nghị táp chí": "2021, AAAI - Xếp hạng A*",
    },

    # ===== 7. Data Maps =====
    {
        "STT": 7,
        "Đề tài / Công bố khoa học": "Dataset Cartography: Mapping and Diagnosing Datasets with Training Dynamics",
        "Loại bài toán": "Phân tích dữ liệu và Lựa chọn tập cốt lõi bằng Động lực huấn luyện",
        "Mục tiêu": "Tạo 'Bản đồ dữ liệu' (Data Map) để chẩn đoán chất lượng từng mẫu trong dataset, xác định mẫu dễ/khó/mơ hồ.",
        "Ngữ cảnh": "Không phải mọi mẫu dữ liệu đều có giá trị như nhau. Một số mẫu dễ học (Easy), một số khó (Hard), một số mơ hồ (Ambiguous - có thể là nhiễu).",
        "Input": "Dữ liệu huấn luyện D, mô hình neural network.",
        "Output": "Bản đồ dữ liệu (Data Map): Mỗi dòng được gán 2 chỉ số Confidence và Variability.\nTập con S được chọn từ vùng mong muốn.",
        "Data": "SNLI, MNLI, WinoGrande, QNLI (các bộ dữ liệu NLP).",
        "Ingest": "Huấn luyện mô hình qua nhiều epoch, ghi lại xác suất dự đoán của từng mẫu ở từng epoch.",
        "Process": "1. Huấn luyện mô hình (VD: RoBERTa) qua E epoch.\n2. Với mỗi mẫu i, tính:\n   - Confidence = trung bình p(y_đúng) qua E epoch.\n   - Variability = độ lệch chuẩn p(y_đúng) qua E epoch.\n3. Vẽ Bản đồ dữ liệu 2D (Confidence vs Variability).\n4. Phân loại: Easy (conf cao, var thấp), Hard (conf thấp, var thấp), Ambiguous (var cao).",
        "Machine Learning": "RoBERTa, BERT (bản gốc NLP). CoreTab đã điều chỉnh cho XGBoost/GBDT.",
        "Kịch bản thực nghiệm": "1. Chia dữ liệu thành 3 vùng, huấn luyện trên từng vùng.\n2. Chỉ giữ vùng Ambiguous → tốt cho khái quát hóa ngoài phân phối (OOD).\n3. Loại bỏ vùng Hard → thường là nhiễu nhãn.",
        "Phương pháp đánh giá": "Accuracy, Khái quát hóa OOD, Tỷ lệ phát hiện nhiễu.",
        "Công nghệ và nền tảng triển khai": "Python, PyTorch, HuggingFace Transformers.",
        "Kết quả": "1. Chỉ cần 33% dữ liệu (vùng Ambiguous) để đạt 99% accuracy.\n2. Vùng Hard thường chứa 80%+ nhiễu nhãn.\n3. CoreTab lấy tư tưởng này để tạo Tabular Datamap.",
        "Source": "https://aclanthology.org/2020.emnlp-main.746/",
        "BibTeX": "@inproceedings{swayamdipta2020dataset,\n  title={Dataset Cartography},\n  author={Swayamdipta, S. et al.},\n  booktitle={EMNLP},\n  year={2020}\n}",
        "Báo cáo / Code": "https://github.com/allenai/cartography",
        "Năm xuất bản tên hội nghị táp chí": "2020, EMNLP - Xếp hạng A",
    },

    # ===== 8. Example Forgetting =====
    {
        "STT": 8,
        "Đề tài / Công bố khoa học": "An Empirical Study of Example Forgetting during Deep Neural Network Learning",
        "Loại bài toán": "Phân tích Động lực huấn luyện và Cắt tỉa dữ liệu",
        "Mục tiêu": "Nghiên cứu hiện tượng 'quên mẫu' (forgetting events) trong quá trình huấn luyện DNN, từ đó xác định mẫu nào có thể bỏ an toàn.",
        "Ngữ cảnh": "Trong quá trình huấn luyện, mô hình có thể đoán đúng mẫu A ở epoch 5 nhưng lại đoán sai ở epoch 6 (một sự kiện quên). Một số mẫu bị quên liên tục → rất quan trọng.",
        "Input": "Dữ liệu huấn luyện D, mô hình DNN.",
        "Output": "Điểm Quên (Forgetting Score) cho từng mẫu. Tập con S = các mẫu có forgetting score cao.",
        "Data": "CIFAR-10, CIFAR-100, Permuted MNIST.",
        "Ingest": "Huấn luyện mô hình, ghi lại accuracy của từng mẫu sau mỗi lần xuất hiện.",
        "Process": "1. Huấn luyện mô hình qua nhiều epoch.\n2. Với mỗi mẫu, đếm số lần chuyển từ đúng → sai (sự kiện quên).\n3. Xếp hạng mẫu theo điểm quên.\n4. Các mẫu không bao giờ bị quên (Unforgettable, score=0) có thể bỏ an toàn.\n5. Các mẫu bị quên nhiều là những mẫu quan trọng nhất.",
        "Machine Learning": "ResNet, VGG, WideResNet.",
        "Kịch bản thực nghiệm": "1. Bỏ 30% mẫu unforgettable → accuracy giảm <1%.\n2. Bỏ 10% mẫu high-forgetting → accuracy giảm mạnh.\n3. Điểm quên ổn định qua các kiến trúc khác nhau.",
        "Phương pháp đánh giá": "Test Accuracy sau khi cắt tỉa, phân phối điểm quên.",
        "Công nghệ và nền tảng triển khai": "Python, PyTorch.",
        "Kết quả": "Có thể bỏ an toàn 30-40% dữ liệu (unforgettable) mà accuracy chỉ giảm <1%. Điểm quên chuyển giao được giữa các kiến trúc.",
        "Source": "https://arxiv.org/abs/1812.05159",
        "BibTeX": "@inproceedings{toneva2019empirical,\n  title={Example Forgetting},\n  author={Toneva, M. et al.},\n  booktitle={ICLR},\n  year={2019}\n}",
        "Báo cáo / Code": "https://github.com/mtoneva/example_forgetting",
        "Năm xuất bản tên hội nghị táp chí": "2019, ICLR - Xếp hạng A* (LƯU Ý: Trước 2020)",
    },

    # ===== 9. SVP =====
    {
        "STT": 9,
        "Đề tài / Công bố khoa học": "Selection via Proxy: Efficient Data Selection for Deep Learning",
        "Loại bài toán": "Lựa chọn dữ liệu bằng Mô hình đại diện (Proxy Model)",
        "Mục tiêu": "Dùng mô hình nhỏ (Proxy) để thay thế mô hình lớn trong quá trình chọn dữ liệu, tăng tốc gấp 10-40 lần.",
        "Ngữ cảnh": "Active Learning và Coreset Selection thường yêu cầu chạy inference/training trên mô hình lớn để chấm điểm, rất chậm.",
        "Input": "Dữ liệu D, Proxy model (nhỏ), Target model (lớn).",
        "Output": "Tập con S được chọn bởi Proxy, dùng huấn luyện Target model.",
        "Data": "CIFAR-10, CIFAR-100, ImageNet, Amazon Review.",
        "Ingest": "Huấn luyện Proxy model (ít epoch/ít layers), tính điểm cho từng mẫu.",
        "Process": "1. Chọn Proxy model (VD: ResNet-18 thay vì ResNet-164).\n2. Huấn luyện Proxy nhanh (ít epochs).\n3. Dùng Proxy để chấm điểm từng mẫu (uncertainty, gradient norm, v.v.).\n4. Chọn top-k mẫu theo điểm của Proxy.\n5. Huấn luyện Target model trên top-k.",
        "Machine Learning": "ResNet-18 (Proxy), ResNet-164 (Target), WideResNet.",
        "Kịch bản thực nghiệm": "1. So sánh SVP vs Active Learning (không proxy) về tốc độ và accuracy.\n2. Nhiều tỷ lệ nén: 10%, 30%, 50%.\n3. Kiểm tra hệ số tương quan thứ hạng giữa Proxy và Target.",
        "Phương pháp đánh giá": "Test Accuracy, Thời gian lựa chọn, Hệ số tương quan Spearman.",
        "Công nghệ và nền tảng triển khai": "Python, PyTorch.",
        "Kết quả": "SVP tăng tốc 41.9 lần cho Active Learning, mất <0.1% accuracy. Hệ số tương quan giữa Proxy và Target rất cao (>0.95).",
        "Source": "https://arxiv.org/abs/1906.11829",
        "BibTeX": "@inproceedings{coleman2020selection,\n  title={Selection via Proxy},\n  author={Coleman, C. et al.},\n  booktitle={ICLR},\n  year={2020}\n}",
        "Báo cáo / Code": "https://github.com/stanford-futuredata/selection-via-proxy",
        "Năm xuất bản tên hội nghị táp chí": "2020, ICLR - Xếp hạng A*",
    },

    # ===== 10. Moderate Coreset =====
    {
        "STT": 10,
        "Đề tài / Công bố khoa học": "Moderate-fitting as a Natural Regularization for Data Pruning",
        "Loại bài toán": "Lựa chọn tập cốt lõi bằng Hình học (khoảng cách trung vị)",
        "Mục tiêu": "Chọn các mẫu nằm ở vùng 'trung dung' trong không gian đặc trưng, tránh chọn mẫu quá dễ hoặc quá khó.",
        "Ngữ cảnh": "Các phương pháp trước thường chọn mẫu khó nhất (hardest) hoặc dễ nhất (easiest). Cả hai cách đều có nhược điểm riêng.",
        "Input": "Dữ liệu D, biểu diễn đặc trưng (feature embeddings).",
        "Output": "Tập con S gồm các mẫu ở khoảng cách trung bình đến tâm của class.",
        "Data": "CIFAR-10, CIFAR-100, ImageNet, Tiny-ImageNet.",
        "Ingest": "Trích xuất biểu diễn đặc trưng từ mô hình đã huấn luyện trước.",
        "Process": "1. Tính biểu diễn đặc trưng cho từng mẫu.\n2. Tính tâm (centroid) của mỗi lớp.\n3. Tính khoảng cách mỗi mẫu đến tâm lớp của nó.\n4. Chọn các mẫu có khoảng cách TRUNG BÌNH (không quá gần, không quá xa).\n5. Xuất coreset.",
        "Machine Learning": "ResNet, VGG (bản gốc là ảnh). Có thể áp dụng cho Tabular bằng cách dùng feature embedding từ MLP hoặc XGBoost leaves.",
        "Kịch bản thực nghiệm": "So sánh với Random, K-Center, Forgetting trên nhiều tỷ lệ nén.",
        "Phương pháp đánh giá": "Test Accuracy, so sánh với các phương pháp chọn Hard-only/Easy-only.",
        "Công nghệ và nền tảng triển khai": "Python, PyTorch.",
        "Kết quả": "Moderate-fitting vượt các phương pháp chọn Hard-only hoặc Easy-only, đặc biệt tốt ở tỷ lệ nén thấp (1-10%).",
        "Source": "https://arxiv.org/abs/2210.01093",
        "BibTeX": "@article{xie2023moderate,\n  title={Moderate-fitting},\n  author={Xie, Y. et al.},\n  journal={arXiv:2210.01093},\n  year={2023}\n}",
        "Báo cáo / Code": "https://github.com/tmllab/Moderate-DS",
        "Năm xuất bản tên hội nghị táp chí": "2023, arXiv (trích dẫn cao, >100 citations)",
    },

    # ===== 11. DC =====
    {
        "STT": 11,
        "Đề tài / Công bố khoa học": "Dataset Condensation with Gradient Matching",
        "Loại bài toán": "Ngưng đọng dữ liệu (Dataset Condensation) - Sinh dữ liệu ảo",
        "Mục tiêu": "Tạo tập dữ liệu ẢO (synthetic) cực nhỏ sao cho mô hình huấn luyện trên tập ảo đạt hiệu năng tương đương mô hình huấn luyện trên tập thật.",
        "Ngữ cảnh": "Khác với Coreset Selection (CHỌN dòng thật), Dataset Condensation TẠO RA dòng mới. Kích thước tập ảo có thể nhỏ hơn coreset rất nhiều (VD: 1 ảnh/class).",
        "Input": "Dữ liệu gốc D.",
        "Output": "Tập dữ liệu ảo S (synthetic), |S| << |D|, các dòng trong S không có trong D.",
        "Data": "MNIST, CIFAR-10, CIFAR-100.",
        "Ingest": "Khởi tạo tập ảo S ngẫu nhiên.",
        "Process": "1. Khởi tạo S ngẫu nhiên.\n2. 'Vòng trong' (Inner loop): Huấn luyện mô hình θ trên S qua T bước.\n3. 'Vòng ngoài' (Outer loop): Cập nhật S sao cho gradient trên S khớp gradient trên D.\n4. Lặp lại cho đến khi S hội tụ.\n5. Xuất S (dữ liệu ảo).",
        "Machine Learning": "ConvNet, ResNet (bản gốc là ảnh).",
        "Kịch bản thực nghiệm": "1. So sánh Accuracy khi huấn luyện trên S vs Random coreset cùng kích thước.\n2. Nhiều IPC (Ảnh trên mỗi lớp): 1, 10, 50.\n3. Chuyển giao kiến trúc: Nén bằng ConvNet, đánh giá bằng ResNet.",
        "Phương pháp đánh giá": "Test Accuracy, so sánh với Random và ngân sách tương đương.",
        "Công nghệ và nền tảng triển khai": "Python, PyTorch.",
        "Kết quả": "DC vượt Random Sampling 10-30% accuracy ở cùng kích thước. Tuy nhiên không phù hợp trực tiếp cho Tabular với nhiều cột Categorical.",
        "Source": "https://arxiv.org/abs/2006.05929",
        "BibTeX": "@inproceedings{zhao2021dataset,\n  title={Dataset Condensation with Gradient Matching},\n  author={Zhao, B. et al.},\n  booktitle={ICLR},\n  year={2021}\n}",
        "Báo cáo / Code": "https://github.com/VICO-UoE/DatasetCondensation",
        "Năm xuất bản tên hội nghị táp chí": "2021, ICLR - Xếp hạng A*",
    },

    # ===== 12. DM =====
    {
        "STT": 12,
        "Đề tài / Công bố khoa học": "Dataset Condensation with Distribution Matching",
        "Loại bài toán": "Ngưng đọng dữ liệu bằng Khớp phân phối",
        "Mục tiêu": "Tạo tập dữ liệu ảo bằng cách khớp phân phối xác suất (thay vì gradient), nhanh hơn DC rất nhiều.",
        "Ngữ cảnh": "DC cần tính gradient bậc 2 (rất tốn kém). DM đơn giản hóa bằng cách chỉ cần khớp phân phối đặc trưng.",
        "Input": "Dữ liệu gốc D.",
        "Output": "Tập dữ liệu ảo S có phân phối tương tự D.",
        "Data": "CIFAR-10, CIFAR-100, Tiny-ImageNet.",
        "Ingest": "Khởi tạo S ngẫu nhiên.",
        "Process": "1. Khởi tạo S ngẫu nhiên.\n2. Cho D và S qua mạng neural để lấy biểu diễn đặc trưng.\n3. Tính Maximum Mean Discrepancy (MMD) giữa embedding của D và S.\n4. Cập nhật S để giảm MMD.\n5. Xuất S.",
        "Machine Learning": "ConvNet, ResNet.",
        "Kịch bản thực nghiệm": "So sánh với DC, DSA, Random trên nhiều IPC.",
        "Phương pháp đánh giá": "Test Accuracy, Thời gian huấn luyện, Bộ nhớ sử dụng.",
        "Công nghệ và nền tảng triển khai": "Python, PyTorch.",
        "Kết quả": "DM nhanh hơn DC 10 lần, accuracy chỉ thấp hơn 1-3%. Phù hợp cho dữ liệu lớn hơn.",
        "Source": "https://arxiv.org/abs/2110.04181",
        "BibTeX": "@inproceedings{zhao2023dm,\n  title={Distribution Matching},\n  author={Zhao, B. and Bilen, H.},\n  booktitle={WACV},\n  year={2023}\n}",
        "Báo cáo / Code": "https://github.com/VICO-UoE/DatasetCondensation",
        "Năm xuất bản tên hội nghị táp chí": "2023, WACV (IEEE)",
    },

    # ===== 13. C2TC =====
    {
        "STT": 13,
        "Đề tài / Công bố khoa học": "C2TC: A Training-Free Framework for Efficient Tabular Data Condensation",
        "Loại bài toán": "Ngưng đọng dữ liệu ĐẶC TRỊ cho bảng (Training-Free Tabular Condensation)",
        "Mục tiêu": "Nén dữ liệu bảng KHÔNG CẦN GRADIENT (training-free), cực nhanh, xử lý được đặc trưng phân loại (Categorical features).",
        "Ngữ cảnh": "DC và DM thiết kế cho ảnh, khi áp dụng cho dữ liệu bảng (có nhiều cột Categorical) thì kém hiệu quả. C2TC giải quyết vấn đề này.",
        "Input": "Dữ liệu bảng D (số + categorical).",
        "Output": "Tập dữ liệu ảo S đặc trị cho Tabular.",
        "Data": "Các bộ dữ liệu tabular thực tế.",
        "Ingest": "Đọc dữ liệu bảng, mã hóa Categorical bằng HCFE (Hybrid Categorical Feature Encoding).",
        "Process": "1. Mã hóa đặc trưng Categorical bằng HCFE.\n2. Định nghĩa bài toán CCAP (Class-Adaptive Cluster Allocation Problem).\n3. Giải CCAP bằng HFILS (Heuristic Local Search): Luân phiên phân bổ mềm và phân cụm theo từng lớp.\n4. Xuất tập S đặc trị tabular.",
        "Machine Learning": "XGBoost, LightGBM, MLP, LogisticRegression (hỗ trợ đa lớp tự nhiên nhờ phân cụm thích nghi theo lớp).",
        "Kịch bản thực nghiệm": "So sánh với DC, DM, Random trên các bộ dữ liệu tabular. Kiểm tra tốc độ (speedup) và accuracy.",
        "Phương pháp đánh giá": "Test Accuracy, Tốc độ tăng tốc (so với DC/DM), hỗ trợ đa lớp.",
        "Công nghệ và nền tảng triển khai": "Python, scikit-learn.",
        "Kết quả": "C2TC nhanh hơn DC/DM 100 lần, giữ nguyên hoặc vượt accuracy trên dữ liệu tabular. Xử lý tốt Categorical features.",
        "Source": "https://arxiv.org/abs/2602.21717",
        "BibTeX": "@inproceedings{c2tc2026,\n  title={C2TC},\n  author={Sara et al.},\n  booktitle={ICDE},\n  year={2026}\n}",
        "Báo cáo / Code": "https://github.com/Sssara-5/TF-TabularCondensation",
        "Năm xuất bản tên hội nghị táp chí": "2026, ICDE (IEEE) - Xếp hạng A*",
    },

    # ===== 14. TDColER =====
    {
        "STT": 14,
        "Đề tài / Công bố khoa học": "TDColER: On Learning Representations for Tabular Data Distillation",
        "Loại bài toán": "Chưng cất dữ liệu bảng bằng Nhúng cột (Column Embedding)",
        "Mục tiêu": "Dùng Nhúng cột (Column Embedding) để học biểu diễn tốt hơn cho dữ liệu bảng trước khi chưng cất, tăng chất lượng 0.5-143%.",
        "Ngữ cảnh": "Các phương pháp chưng cất trước áp dụng trực tiếp cho Tabular mà không xem xét đặc tính riêng (tính không đồng nhất của đặc trưng, categorical vs numerical).",
        "Input": "Dữ liệu bảng D (số + categorical).",
        "Output": "Tập dữ liệu ảo S chất lượng cao hơn nhờ Column Embedding.",
        "Data": "23 bộ dữ liệu tabular (TDBench benchmark).",
        "Ingest": "Đọc dữ liệu bảng, dùng Column Embedding để mã hóa từng cột.",
        "Process": "1. Học Nhúng cột (Column Embedding) cho mỗi cột của D.\n2. Dùng kiến trúc bộ mã hóa-giải mã (Encoder-Decoder) để tạo biểu diễn tiềm ẩn.\n3. Áp dụng phương pháp chưng cất (DC, DM, v.v.) trên biểu diễn mới.\n4. Giải mã về không gian gốc.\n5. Xuất tập S.",
        "Machine Learning": "7 nhóm mô hình: XGBoost, LightGBM, CatBoost, MLP, LogReg, KNN, Decision Tree.",
        "Kịch bản thực nghiệm": "Benchmark TDBench: 23 datasets × 7 models × 11 phương pháp chưng cất. So sánh TDColER vs phương pháp gốc.",
        "Phương pháp đánh giá": "Test Accuracy, Hiệu suất tương đối (so với không có TDColER).",
        "Công nghệ và nền tảng triển khai": "Python, PyTorch, scikit-learn.",
        "Kết quả": "TDColER tăng chất lượng chưng cất 0.5-143% trên toàn bộ TDBench.",
        "Source": "ArXiv 2025 (kèm TDBench paper)",
        "BibTeX": "@article{kang2025tdcoler,\n  title={TDColER},\n  author={Kang, Inwon et al.},\n  journal={arXiv},\n  year={2025}\n}",
        "Báo cáo / Code": "https://github.com/inwonakng/tdbench",
        "Năm xuất bản tên hội nghị táp chí": "2025, ArXiv (kèm TDBench benchmark)",
    },

    # ===== 15. Random =====
    {
        "STT": 15,
        "Đề tài / Công bố khoa học": "Lấy mẫu Ngẫu nhiên Đều (Uniform Random Sampling - RAN) - Phương pháp cơ sở",
        "Loại bài toán": "Phương pháp cơ sở (Baseline - Mốc tham chiếu)",
        "Mục tiêu": "Lấy mẫu hoàn toàn ngẫu nhiên từ tập huấn luyện, làm mốc để so sánh với các phương pháp nén thông minh hơn.",
        "Ngữ cảnh": "Mọi nghiên cứu đánh giá chuẩn (benchmarking) nghiêm túc đều cần một Baseline đơn giản để chứng minh các phương pháp mới có giá trị. CoreTab dùng RAN làm Baseline số 1.",
        "Input": "Dữ liệu huấn luyện D, tỷ lệ nén r.",
        "Output": "Tập con S có kích thước |S| = r × |D|, chọn ngẫu nhiên.",
        "Data": "Mọi bộ dữ liệu.",
        "Ingest": "Đọc dữ liệu.",
        "Process": "numpy.random.choice(n, size=int(n*r), replace=False)\nHoàn toàn ngẫu nhiên, không có thuật toán phức tạp.",
        "Machine Learning": "Mọi mô hình.",
        "Kịch bản thực nghiệm": "Chạy song song với 14 phương pháp còn lại, làm Giới hạn dưới (Lower Bound).",
        "Phương pháp đánh giá": "Cùng các chỉ số: F1-Macro, AUC-ROC, Thời gian.",
        "Công nghệ và nền tảng triển khai": "numpy, scikit-learn.",
        "Kết quả": "Thường là phương pháp kém nhất (Giới hạn dưới). Bất kỳ phương pháp nào không thắng được RAN thì vô giá trị.",
        "Source": "N/A (Phương pháp cơ sở chuẩn)",
        "BibTeX": "N/A",
        "Báo cáo / Code": "numpy.random.choice\nsklearn.utils.resample",
        "Năm xuất bản tên hội nghị táp chí": "N/A (Phương pháp cơ sở chuẩn)",
    },
]

df = pd.DataFrame(data)

output_path = r"C:\KLTN\paper\Plan (2).xlsx"
with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name="Khảo sát tổng quan")
    ws = writer.sheets["Khảo sát tổng quan"]

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    widths = {
        "A": 5, "B": 50, "C": 30, "D": 45, "E": 40,
        "F": 35, "G": 35, "H": 45, "I": 35, "J": 55,
        "K": 35, "L": 50, "M": 40, "N": 30, "O": 50,
        "P": 40, "Q": 40, "R": 40, "S": 30,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    for i in range(2, ws.max_row + 1):
        ws.row_dimensions[i].height = 200

print("DONE: Plan(2).xlsx - Ban tieng Viet CO DAU da tao thanh cong!")
print(f"So dong: {len(data)}, So cot: {len(df.columns)}")
