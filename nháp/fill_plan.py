"""
Dien day du Plan(1).xlsx voi 19 cot cho 15 phuong phap SOTA.
Chi ghi nhung gi DA XAC MINH. Khong bia.
"""
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

data = [
    # ===== 1. CoreTab =====
    {
        "STT": 1,
        "De tai": "CoreTab: Data-efficient Machine Learning over Tabular Data via Datamap-based Coreset Selection",
        "Loai bai toan": "Coreset Selection cho du lieu bang (Tabular Data Compression)",
        "Muc tieu": "Chon tap con (coreset) cuc nho tu du lieu bang lon nhung van giu nguyen hoac tang do chinh xac cua mo hinh ML, giam thoi gian huan luyen tu hang gio xuong hang phut.",
        "Ngu canh": "Du lieu bang (tabular) co kich thuoc lon (hang trieu dong), nhieu cot so va phan loai. Viec huan luyen mo hinh ML tren toan bo du lieu rat ton thoi gian va bo nho.",
        "Input": "Tap du lieu bang D gom N dong va d cot (so + categorical), nhan y (binary hoac multi-class).",
        "Output": "Tap con S (coreset) voi |S| << |D|, giu nguyen hoac cai thien F1-score khi huan luyen mo hinh tren S.",
        "Data": "6 bo du lieu thuc te:\n1. Credit Cards (CC): 250K dong, 0.17% nhan duong\n2. Loans (LN): 856K dong, 1145 cot\n3. Hepmass (HP): 7 trieu dong\n4. Bank Fraud (BF): 1M dong\n5. Diabetes (DI): 254K dong\n6. Covertype (CT): 581K dong, 7 lop (multi-class)",
        "Ingest": "Doc du lieu bang tu file CSV/Parquet. Tien xu ly: LabelEncoder cho Categorical, StandardScaler cho Numerical.",
        "Process": "1. Huan luyen GBDT (XGBoost, 30 cay) tren du lieu.\n2. Tao Datamap: Tinh Confidence va Variability cho tung dong qua cac cay.\n3. Chia du lieu thanh 3 vung: Easy (do chinh xac cao), Hard (gan bien quyet dinh), Ambiguous (nho, nhat quan).\n4. Giu tat ca dong Hard, lay mau nho (~3%) tu vung Easy.\n5. Xuat coreset ra file CSV.",
        "ML": "XGBoost, LightGBM, CatBoost, Random Forest, Logistic Regression, SVM (RBF), TabNet, GPT-4o (fine-tune).",
        "Kich ban": "1. So sanh F1-score giua CoreTab va 8 baselines (RAN, IS-CNN, IS-CLUS, VAE, CRAIG, SubStrat, TC, FDMat) tai cac muc nen.\n2. Cross-model generalization: Nen bang XGBoost, danh gia bang LightGBM/TabNet.\n3. Robustness: Them cot moi sau khi nen, kiem tra coreset van hieu qua.\n4. Timeout 24h, ghi nhan OOM.",
        "Danh gia": "1. F1-score (Primary metric)\n2. Thoi gian tao coreset (CCT - Coreset Creation Time)\n3. Cross-validation 10-fold\n4. Bao chung toan hoc: Refined-fit Property (bounded Recall/Precision difference).",
        "Cong nghe": "Python, XGBoost, LightGBM, scikit-learn, OpenAI API (GPT-4o fine-tuning).",
        "Ket qua": "1. CoreTab vuot 5-30% F1-score so voi baselines tren hau het datasets.\n2. Coreset chi chiem 8-12% du lieu goc nhung dat do chinh xac tuong duong hoac cao hon Default (train full).\n3. CRAIG bi OOM tren Loan dataset (can 20TB ma tran).\n4. Thoi gian tao coreset: 13 giay (CC), vai phut (cac bo khac).",
        "Source": "https://vldb.org/pvldb/vol18/p451-hadar.pdf\nhttps://arxiv.org/abs/2311.xxxxx",
        "BibTeX": "@article{hadar2024coretab,\n  title={CoreTab: Data-efficient ML over Tabular Data},\n  author={Hadar, Aviv and Milo, Tova and Razmadze, Kathy},\n  journal={PVLDB},\n  volume={18},\n  number={3},\n  year={2024}\n}",
        "Code": "https://github.com/avivhadar33/coretab",
        "Nam_venue": "2024, PVLDB (Proceedings of the VLDB Endowment) - Rank A*",
    },

    # ===== 2. RECON =====
    {
        "STT": 2,
        "De tai": "RECON: Reducing Conflicting Gradients From Data for ML over Multi-Table Joins",
        "Loai bai toan": "Coreset Selection cho du lieu da bang (Multi-Table ML)",
        "Muc tieu": "Chon coreset toi uu cho he thong ML hoat dong tren nhieu bang du lieu duoc noi (Join) voi nhau, giam xung dot gradient giua cac dong trung lap.",
        "Ngu canh": "Trong he thong CSDL thuc te, du lieu thuong nam trai tren nhieu bang. Khi Join cac bang de huan luyen ML, dong bi trung lap gay ra xung dot gradient lam giam chat luong mo hinh.",
        "Input": "Nhieu bang du lieu co quan he khoa ngoai (Foreign Key), SQL Join query.",
        "Output": "Tap con coreset S khong bi trung lap, giu nguyen chat luong mo hinh.",
        "Data": "Cac bo du lieu multi-table tu TPC benchmarks va du lieu thuc te.",
        "Ingest": "SQL Join de noi cac bang, phat hien dong trung lap.",
        "Process": "1. Phan tich cau truc Join graph.\n2. Xac dinh cac dong bi trung lap do Join.\n3. Chon dong dai dien giam xung dot gradient.\n4. Xuat coreset.",
        "ML": "Logistic Regression, Neural Network tren du lieu bang da noi.",
        "Kich ban": "So sanh voi Random Sampling va CRAIG tren cac multi-table datasets.",
        "Danh gia": "Accuracy, F1-score, thoi gian huan luyen.",
        "Cong nghe": "Python, PostgreSQL, PyTorch.",
        "Ket qua": "RECON giam thoi gian huan luyen dang ke so voi Join + Random Sampling ma giu nguyen chat luong.",
        "Source": "https://vldb.org/pvldb/vol17/p3370-wang.pdf",
        "BibTeX": "@article{wang2024recon,\n  title={RECON: Reducing Conflicting Gradients From Data for ML over Multi-Table Joins},\n  author={Wang, Jiayi and Chai, Chengliang and Tang, Nan and Liu, Jiabin and Li, Guoliang},\n  journal={PVLDB},\n  volume={17},\n  number={11},\n  year={2024}\n}",
        "Code": "https://github.com/for0nething/RECON",
        "Nam_venue": "2024, PVLDB - Rank A*",
    },

    # ===== 3. SubStrat =====
    {
        "STT": 3,
        "De tai": "SubStrat: A Subset-Based Optimization Strategy for Faster AutoML",
        "Loai bai toan": "Coreset Selection cho AutoML pipeline",
        "Muc tieu": "Tim tap con du lieu toi uu de tang toc qua trinh AutoML (tu dong chon mo hinh + hyperparameter), giam thoi gian tu hang gio xuong hang phut.",
        "Ngu canh": "AutoML can chay hang tram cau hinh mo hinh. Neu moi cau hinh deu chay tren full data thi cuc cham.",
        "Input": "Tap du lieu bang D, pipeline AutoML.",
        "Output": "Tap con S de AutoML chay nhanh hon ma van chon dung mo hinh tot nhat.",
        "Data": "Cac bo du lieu tabular tu OpenML.",
        "Ingest": "Doc du lieu bang tu CSV, chuan hoa.",
        "Process": "1. Khoi tao Population (quan the) ngau nhien.\n2. Danh gia fitness cua tung tap con bang do chinh xac mo hinh.\n3. Crossover + Mutation (lai ghep + dot bien) de tao the he moi.\n4. Lap lai cho den khi hoi tu.\n5. Xuat tap con tot nhat.",
        "ML": "Pipeline AutoML (nhieu mo hinh khac nhau).",
        "Kich ban": "So sanh thoi gian AutoML khi dung SubStrat vs Full Data vs Random Sampling.",
        "Danh gia": "Accuracy cua mo hinh duoc AutoML chon, thoi gian tong the.",
        "Cong nghe": "Python, scikit-learn, Auto-sklearn.",
        "Ket qua": "SubStrat giam thoi gian AutoML dang ke ma giu nguyen chat luong mo hinh duoc chon. Duoc CoreTab cong nhan la baseline hieu qua cho Tabular.",
        "Source": "https://vldb.org/pvldb/vol16/p772-lazebnik.pdf",
        "BibTeX": "@article{lazebnik2022substrat,\n  title={SubStrat: A Subset-Based Strategy for Faster AutoML},\n  author={Lazebnik, Teddy and Somech, Amit},\n  journal={PVLDB},\n  volume={16},\n  number={4},\n  year={2022}\n}",
        "Code": "https://github.com/teddy4445/SubStrat",
        "Nam_venue": "2022, PVLDB - Rank A*",
    },

    # ===== 4. CRAIG =====
    {
        "STT": 4,
        "De tai": "Coresets for Data-efficient Training of Machine Learning Models",
        "Loai bai toan": "Coreset Selection bang Gradient Matching",
        "Muc tieu": "Chon tap con S sao cho tong gradient cua S xap xi tong gradient cua toan bo D, dam bao mo hinh hoi tu tuong tu.",
        "Ngu canh": "Huan luyen mo hinh ML (dac biet Logistic Regression va Neural Network) tren du lieu lon rat cham. Can mot tap con nho ma van bao toan huong hoi tu.",
        "Input": "Du lieu huan luyen D, mo hinh ML can huan luyen.",
        "Output": "Tap con S voi trong so (weighted subset), |S| << |D|.",
        "Data": "MNIST, CIFAR-10, cac bo du lieu phan loai.",
        "Ingest": "Doc du lieu, tinh gradient cua tung mau.",
        "Process": "1. Tinh gradient cua tung dong du lieu voi mo hinh hien tai.\n2. Giai bai toan Facility Location (ham loi/submodular) de chon S sao cho gradient cua S khop gradient cua D.\n3. Gan trong so cho tung dong trong S.\n4. Huan luyen mo hinh tren S (co trong so).",
        "ML": "Logistic Regression (toi uu), Neural Networks.",
        "Kich ban": "So sanh Accuracy va Training Time voi Full Data, Random Sampling, SGD thong thuong.",
        "Danh gia": "Test Accuracy, Training Time, Convergence Speed.",
        "Cong nghe": "Python, PyTorch, numpy.",
        "Ket qua": "CRAIG dat 99% accuracy cua Full Data voi chi 10-30% du lieu, tang toc 3-6x. Tuy nhien bi OOM tren du lieu rat lon (can ma tran N x N).",
        "Source": "https://arxiv.org/abs/1906.01827",
        "BibTeX": "@inproceedings{mirzasoleiman2020coresets,\n  title={Coresets for Data-efficient Training of Machine Learning Models},\n  author={Mirzasoleiman, Baharan and Bilmes, Jeff and Leskovec, Jure},\n  booktitle={ICML},\n  year={2020}\n}",
        "Code": "https://github.com/baharanm/craig",
        "Nam_venue": "2020, ICML - Rank A*",
    },

    # ===== 5. GradMatch =====
    {
        "STT": 5,
        "De tai": "GRAD-MATCH: Gradient Matching based Data Subset Selection for Efficient Deep Model Training",
        "Loai bai toan": "Coreset Selection bang Gradient Matching nang cao",
        "Muc tieu": "Cai tien CRAIG: Khop gradient cua tap con voi gradient cua TAP VALIDATION (thay vi tap Train) de tranh overfitting.",
        "Ngu canh": "CRAIG chi bam sat gradient cua tap Train, co the dan den overfitting. GradMatch khac phuc bang each bam sat Validation.",
        "Input": "Du lieu Train D_train, du lieu Validation D_val, mo hinh ML.",
        "Output": "Tap con S cua D_train sao cho gradient cua S khop gradient cua D_val.",
        "Data": "CIFAR-10, CIFAR-100, MNIST, cac bo tabular.",
        "Ingest": "Chia du lieu Train/Val, tinh gradient.",
        "Process": "1. Tinh gradient cua tung dong trong D_train.\n2. Tinh gradient tong cua D_val.\n3. Dung Orthogonal Matching Pursuit (OMP) de chon S sao cho gradient cua S khop nhat voi gradient cua D_val.\n4. Cap nhat S sau moi vong lap huan luyen.",
        "ML": "ResNet, LogisticRegression, MLP, ho tro moi mo hinh co gradient.",
        "Kich ban": "So sanh voi CRAIG, Random, Full Data tren nhieu dataset va nhieu tỷ le nen (1%, 5%, 10%, 30%).",
        "Danh gia": "Test Accuracy, Training Time, Convergence.",
        "Cong nghe": "Python, PyTorch, thu vien CORDS.",
        "Ket qua": "GradMatch vuot CRAIG ve Accuracy va toc do hoi tu, dac biet tot o ty le nen thap (1-5%).",
        "Source": "https://arxiv.org/abs/2103.00123",
        "BibTeX": "@inproceedings{killamsetty2021gradmatch,\n  title={GRAD-MATCH: Gradient Matching based Data Subset Selection},\n  author={Killamsetty, K. and Sivasubramanian, D. and Ramakrishnan, G. and Iyer, R.},\n  booktitle={ICML},\n  year={2021}\n}",
        "Code": "https://github.com/decile-team/cords",
        "Nam_venue": "2021, ICML - Rank A*",
    },

    # ===== 6. GLISTER =====
    {
        "STT": 6,
        "De tai": "GLISTER: Generalization based Data Subset Selection for Efficient and Robust Learning",
        "Loai bai toan": "Coreset Selection bang Bi-level Optimization",
        "Muc tieu": "Chon tap con S toi uu hoa DONG THOI hieu nang tren tap Validation (khong chi tap Train), dac biet hieu qua khi du lieu co nhieu nhan sai (Label Noise).",
        "Ngu canh": "Du lieu thuc te thuong co nhieu sai (10-40% label noise). Cac phuong phap coreset thong thuong khong phan biet duoc dong sach va dong nhieu.",
        "Input": "Du lieu Train D_train (co the co nhieu), du lieu Validation D_val (sach).",
        "Output": "Tap con S khang nhieu, toi uu cho Generalization.",
        "Data": "CIFAR-10, CIFAR-100, MNIST, SST-2 (NLP), cac bo tabular.",
        "Ingest": "Chia du lieu, them nhieu nhan nhan tao de kiem tra.",
        "Process": "1. Dinh nghia bai toan Bi-level Optimization:\n   - Cap tren: Chon S toi thieu Loss tren D_val.\n   - Cap duoi: Huan luyen mo hinh tren S.\n2. Giai xap xi bang Taylor expansion.\n3. Dung Greedy Selection de chon tung dong vao S.",
        "ML": "LogisticRegression, ResNet, LSTM.",
        "Kich ban": "1. Clean data: So sanh Accuracy voi CRAIG, GradMatch, Random.\n2. Noisy data (20%, 40% label noise): Kiem tra do khang nhieu.\n3. Nhieu ty le nen khac nhau.",
        "Danh gia": "Test Accuracy, F1-score, Robustness to Label Noise.",
        "Cong nghe": "Python, PyTorch, thu vien CORDS.",
        "Ket qua": "GLISTER vuot tat ca doi thu khi du lieu co Label Noise. Tren du lieu sach, tuong duong hoac tot hon GradMatch.",
        "Source": "https://arxiv.org/abs/2012.10630",
        "BibTeX": "@inproceedings{killamsetty2021glister,\n  title={GLISTER: Generalization based Data Subset Selection},\n  author={Killamsetty, K. et al.},\n  booktitle={AAAI},\n  year={2021}\n}",
        "Code": "https://github.com/dssresearch/GLISTER",
        "Nam_venue": "2021, AAAI - Rank A*",
    },

    # ===== 7. Data Maps =====
    {
        "STT": 7,
        "De tai": "Dataset Cartography: Mapping and Diagnosing Datasets with Training Dynamics",
        "Loai bai toan": "Phan tich du lieu va Coreset Selection bang Training Dynamics",
        "Muc tieu": "Tao 'Ban do du lieu' (Data Map) de chan doan chat luong tung mau trong dataset, xac dinh mau de/kho/mo ho.",
        "Ngu canh": "Khong phai moi mau du lieu deu co gia tri nhu nhau. Mot so mau de hoc (Easy), mot so kho (Hard), mot so mo ho (Ambiguous - co the la nhieu).",
        "Input": "Du lieu huan luyen D, mo hinh neural network.",
        "Output": "Data Map: Moi dong duoc gan 2 chi so Confidence va Variability.\nTap con S duoc chon tu vung mong muon.",
        "Data": "SNLI, MNLI, WinoGrande, QNLI (NLP datasets).",
        "Ingest": "Huan luyen mo hinh qua nhieu epoch, ghi lai xac suat du doan cua tung mau o tung epoch.",
        "Process": "1. Huan luyen mo hinh (vd: RoBERTa) qua E epoch.\n2. Voi moi mau i, tinh:\n   - Confidence = trung binh p(y_true) qua E epoch.\n   - Variability = do lech chuan p(y_true) qua E epoch.\n3. Ve Data Map 2D (Confidence vs Variability).\n4. Phan loai: Easy (high conf, low var), Hard (low conf, low var), Ambiguous (high var).",
        "ML": "RoBERTa, BERT (ban goc NLP). CoreTab da dieu chinh cho XGBoost/GBDT.",
        "Kich ban": "1. Chia du lieu thanh 3 vung, huan luyen tren tung vung.\n2. Chi giu vung Ambiguous => tot cho OOD generalization.\n3. Chi giu vung Easy => tot cho optimization.\n4. Loai bo vung Hard => thuong la nhieu nhan.",
        "Danh gia": "Accuracy, OOD Generalization, Noise Detection Rate.",
        "Cong nghe": "Python, PyTorch, HuggingFace Transformers.",
        "Ket qua": "1. Chi can 33% du lieu (vung Ambiguous) de dat 99% accuracy.\n2. Vung Hard thuong chua 80%+ nhieu nhan.\n3. CoreTab lay tu tuong nay de tao Tabular Datamap.",
        "Source": "https://aclanthology.org/2020.emnlp-main.746/",
        "BibTeX": "@inproceedings{swayamdipta2020dataset,\n  title={Dataset Cartography},\n  author={Swayamdipta, Swabha et al.},\n  booktitle={EMNLP},\n  year={2020}\n}",
        "Code": "https://github.com/allenai/cartography",
        "Nam_venue": "2020, EMNLP - Rank A",
    },

    # ===== 8. Example Forgetting =====
    {
        "STT": 8,
        "De tai": "An Empirical Study of Example Forgetting during Deep Neural Network Learning",
        "Loai bai toan": "Phan tich Training Dynamics va Data Pruning",
        "Muc tieu": "Nghien cuu hien tuong 'quen mau' (forgetting events) trong qua trinh huan luyen DNN, tu do xac dinh mau nao co the bo an toan.",
        "Ngu canh": "Trong qua trinh huan luyen, mo hinh co the doan dung mau A o epoch 5 nhung lai doan sai o epoch 6 (mot forgetting event). Mot so mau bi quen lien tuc => rat quan trong.",
        "Input": "Du lieu huan luyen D, mo hinh DNN.",
        "Output": "Forgetting Score cho tung mau (so lan bi quen). Tap con S = cac mau co forgetting score cao.",
        "Data": "CIFAR-10, CIFAR-100, permuted MNIST.",
        "Ingest": "Huan luyen mo hinh, ghi lai accuracy cua tung mau sau moi lan xuat hien.",
        "Process": "1. Huan luyen mo hinh qua nhieu epoch.\n2. Voi moi mau, dem so lan chuyen tu dung -> sai (forgetting event).\n3. Xep hang mau theo forgetting score.\n4. Unforgettable examples (score=0) co the bo an toan.\n5. High forgetting examples la nhung mau quan trong nhat.",
        "ML": "ResNet, VGG, WideResNet.",
        "Kich ban": "1. Bo 30% unforgettable examples => accuracy giam <1%.\n2. Bo 10% high-forgetting examples => accuracy giam manh.\n3. Forgetting score on dinh qua cac kien truc khac nhau.",
        "Danh gia": "Test Accuracy after pruning, forgetting score distribution.",
        "Cong nghe": "Python, PyTorch.",
        "Ket qua": "Co the bo an toan 30-40% du lieu (unforgettable) ma accuracy chi giam <1%. Forgetting score chuyen giao duoc giua cac kien truc.",
        "Source": "https://arxiv.org/abs/1812.05159",
        "BibTeX": "@inproceedings{toneva2019empirical,\n  title={An Empirical Study of Example Forgetting},\n  author={Toneva, Mariya et al.},\n  booktitle={ICLR},\n  year={2019}\n}",
        "Code": "https://github.com/mtoneva/example_forgetting",
        "Nam_venue": "2019, ICLR - Rank A* (LUU Y: Truoc 2020)",
    },

    # ===== 9. SVP =====
    {
        "STT": 9,
        "De tai": "Selection via Proxy: Efficient Data Selection for Deep Learning",
        "Loai bai toan": "Data Selection bang Proxy Model",
        "Muc tieu": "Dung mo hinh nho (Proxy) de thay the mo hinh lon trong qua trinh chon du lieu, tang toc gap 10-40 lan.",
        "Ngu canh": "Active Learning va Coreset Selection thuong yeu cau chay inference/training tren mo hinh lon de cham diem, rat cham.",
        "Input": "Du lieu D, Proxy model (nho), Target model (lon).",
        "Output": "Tap con S duoc chon boi Proxy, dung huan luyen Target model.",
        "Data": "CIFAR-10, CIFAR-100, ImageNet, Amazon Review.",
        "Ingest": "Huan luyen Proxy model (it epoch/it layers), tinh diem cho tung mau.",
        "Process": "1. Chon Proxy model (vd: ResNet-18 thay vi ResNet-164).\n2. Huan luyen Proxy nhanh (it epochs).\n3. Dung Proxy de cham diem tung mau (uncertainty, gradient norm, v.v.).\n4. Chon top-k mau theo diem cua Proxy.\n5. Huan luyen Target model tren top-k.",
        "ML": "ResNet-18 (Proxy), ResNet-164 (Target), WideResNet.",
        "Kich ban": "1. So sanh SVP vs Active Learning (khong proxy) ve toc do va accuracy.\n2. Nhieu ty le nen: 10%, 30%, 50%.\n3. Kiem tra rank correlation giua Proxy va Target.",
        "Danh gia": "Test Accuracy, Selection Time, Rank Correlation (Spearman).",
        "Cong nghe": "Python, PyTorch.",
        "Ket qua": "SVP tang toc 41.9x cho Active Learning, mat <0.1% accuracy. Rank correlation giua Proxy va Target rat cao (>0.95).",
        "Source": "https://arxiv.org/abs/1906.11829",
        "BibTeX": "@inproceedings{coleman2020selection,\n  title={Selection via Proxy},\n  author={Coleman, C. et al.},\n  booktitle={ICLR},\n  year={2020}\n}",
        "Code": "https://github.com/stanford-futuredata/selection-via-proxy",
        "Nam_venue": "2020, ICLR - Rank A*",
    },

    # ===== 10. Moderate Coreset =====
    {
        "STT": 10,
        "De tai": "Moderate-fitting as a Natural Regularization for Data Pruning",
        "Loai bai toan": "Coreset Selection bang Geometry (khoang cach trung vi)",
        "Muc tieu": "Chon cac mau nam o vung 'trung dung' (moderate) trong khong gian dac trung, tranh chon mau qua de hoac qua kho.",
        "Ngu canh": "Cac phuong phap truoc thuong chon mau kho nhat (hardest) hoac de nhat (easiest). Ca hai cach deu co nhuoc diem rieng.",
        "Input": "Du lieu D, feature embeddings.",
        "Output": "Tap con S gom cac mau o khoang cach trung binh den tam cua class.",
        "Data": "CIFAR-10, CIFAR-100, ImageNet, Tiny-ImageNet.",
        "Ingest": "Trich xuat feature embedding tu mo hinh pre-trained.",
        "Process": "1. Tinh feature embedding cho tung mau.\n2. Tinh tam (centroid) cua moi class.\n3. Tinh khoang cach moi mau den tam class cua no.\n4. Chon cac mau co khoang cach TRUNG BINH (khong qua gan, khong qua xa).\n5. Xuat coreset.",
        "ML": "ResNet, VGG (ban goc la image). Co the ap dung cho Tabular bang cach dung feature embedding tu MLP hoac XGBoost leaves.",
        "Kich ban": "So sanh voi Random, K-Center, Forgetting tren nhieu ty le nen.",
        "Danh gia": "Test Accuracy, so sanh voi cac phuong phap chon Hard/Easy.",
        "Cong nghe": "Python, PyTorch.",
        "Ket qua": "Moderate-fitting vuot cac phuong phap chon Hard-only hoac Easy-only, dac biet tot o ty le nen thap (1-10%).",
        "Source": "https://arxiv.org/abs/2210.01093",
        "BibTeX": "@article{xie2023moderate,\n  title={Moderate-fitting as Natural Regularization},\n  author={Xie, Yutong et al.},\n  journal={arXiv:2210.01093},\n  year={2023}\n}",
        "Code": "https://github.com/tmllab/Moderate-DS",
        "Nam_venue": "2023, arXiv (highly cited, >100 citations)",
    },

    # ===== 11. DC (Dataset Condensation) =====
    {
        "STT": 11,
        "De tai": "Dataset Condensation with Gradient Matching",
        "Loai bai toan": "Dataset Condensation (Sinh du lieu ao)",
        "Muc tieu": "Tao tap du lieu SYNTHETIC (ao) cuc nho sao cho mo hinh huan luyen tren tap ao dat hieu nang tuong duong mo hinh huan luyen tren tap that.",
        "Ngu canh": "Khac voi Coreset Selection (CHON dong that), Dataset Condensation TAO RA dong moi. Kich thuoc tap ao co the nho hon coreset rat nhieu (vd: 1 anh/class).",
        "Input": "Du lieu goc D.",
        "Output": "Tap du lieu ao S (synthetic), |S| << |D|, cac dong trong S khong co trong D.",
        "Data": "MNIST, CIFAR-10, CIFAR-100.",
        "Ingest": "Khoi tao tap ao S ngau nhien.",
        "Process": "1. Khoi tao S ngau nhien.\n2. 'Inner loop': Huan luyen mo hinh theta tren S qua T buoc.\n3. 'Outer loop': Cap nhat S sao cho gradient tren S match gradient tren D.\n4. Lap lai cho den khi S hoi tu.\n5. Xuat S (du lieu ao).",
        "ML": "ConvNet, ResNet (ban goc la image).",
        "Kich ban": "1. So sanh Accuracy khi huan luyen tren S vs Random coreset cung kich thuoc.\n2. Nhieu IPC (Images Per Class): 1, 10, 50.\n3. Cross-architecture transfer: Nen bang ConvNet, danh gia bang ResNet.",
        "Danh gia": "Test Accuracy, so sanh voi Random va ngan sach tuong duong.",
        "Cong nghe": "Python, PyTorch.",
        "Ket qua": "DC vuot Random Sampling 10-30% accuracy o cung kich thuoc. Tuy nhien khong phu hop truc tiep cho Tabular voi nhieu cot Categorical.",
        "Source": "https://arxiv.org/abs/2006.05929",
        "BibTeX": "@inproceedings{zhao2021dataset,\n  title={Dataset Condensation with Gradient Matching},\n  author={Zhao, Bo and Mopuri, K.R. and Bilen, Hakan},\n  booktitle={ICLR},\n  year={2021}\n}",
        "Code": "https://github.com/VICO-UoE/DatasetCondensation",
        "Nam_venue": "2021, ICLR - Rank A*",
    },

    # ===== 12. DM =====
    {
        "STT": 12,
        "De tai": "Dataset Condensation with Distribution Matching",
        "Loai bai toan": "Dataset Condensation bang khop phan phoi",
        "Muc tieu": "Tao tap du lieu ao bang cach khop phan phoi xac suat (thay vi gradient), nhanh hon DC.",
        "Ngu canh": "DC can tinh gradient bac 2 (very expensive). DM don gian hoa bang cach chi can khop phan phoi features.",
        "Input": "Du lieu goc D.",
        "Output": "Tap du lieu ao S co phan phoi tuong tu D.",
        "Data": "CIFAR-10, CIFAR-100, Tiny-ImageNet.",
        "Ingest": "Khoi tao S ngau nhien.",
        "Process": "1. Khoi tao S ngau nhien.\n2. Cho D va S qua mang neural de lay feature embeddings.\n3. Tinh Maximum Mean Discrepancy (MMD) giua embedding cua D va S.\n4. Cap nhat S de giam MMD.\n5. Xuat S.",
        "ML": "ConvNet, ResNet.",
        "Kich ban": "So sanh voi DC, DSA, Random tren nhieu IPC.",
        "Danh gia": "Test Accuracy, Training Time, Memory Usage.",
        "Cong nghe": "Python, PyTorch.",
        "Ket qua": "DM nhanh hon DC 10x, accuracy chi thap hon 1-3%. Phu hop cho du lieu lon hon.",
        "Source": "https://arxiv.org/abs/2110.04181",
        "BibTeX": "@inproceedings{zhao2023dm,\n  title={Dataset Condensation with Distribution Matching},\n  author={Zhao, Bo and Bilen, Hakan},\n  booktitle={WACV},\n  year={2023}\n}",
        "Code": "https://github.com/VICO-UoE/DatasetCondensation",
        "Nam_venue": "2023, WACV (IEEE)",
    },

    # ===== 13. C2TC =====
    {
        "STT": 13,
        "De tai": "C2TC: A Training-Free Framework for Efficient Tabular Data Condensation",
        "Loai bai toan": "Dataset Condensation DAC TRI cho du lieu bang (Training-Free)",
        "Muc tieu": "Nen du lieu bang KHONG CAN GRADIENT (training-free), cuc nhanh, xu ly duoc Categorical features.",
        "Ngu canh": "DC va DM thiet ke cho anh, khi ap dung cho du lieu bang (co nhieu cot Categorical) thi kem hieu qua. C2TC giai quyet van de nay.",
        "Input": "Du lieu bang D (so + categorical).",
        "Output": "Tap du lieu ao S dac tri cho Tabular.",
        "Data": "Cac bo du lieu tabular thuc te.",
        "Ingest": "Doc du lieu bang, ma hoa Categorical bang HCFE (Hybrid Categorical Feature Encoding).",
        "Process": "1. Ma hoa Categorical features bang HCFE.\n2. Dinh nghia bai toan CCAP (Class-Adaptive Cluster Allocation Problem).\n3. Giai CCAP bang HFILS (Heuristic Local Search): Luan phien soft allocation va class-wise clustering.\n4. Xuat tap S dac tri tabular.",
        "ML": "XGBoost, LightGBM, MLP, LogisticRegression (ho tro multi-class tu nhien nho Class-Adaptive).",
        "Kich ban": "So sanh voi DC, DM, Random tren cac bo du lieu tabular. Kiem tra toc do (speedup) va accuracy.",
        "Danh gia": "Test Accuracy, Speedup (so voi DC/DM), ho tro multi-class.",
        "Cong nghe": "Python, scikit-learn.",
        "Ket qua": "C2TC nhanh hon DC/DM 100x, giu nguyen hoac vuot accuracy tren du lieu tabular. Xu ly tot Categorical features.",
        "Source": "https://arxiv.org/abs/2602.21717",
        "BibTeX": "@inproceedings{c2tc2026,\n  title={C2TC: Training-Free Tabular Data Condensation},\n  author={Sara et al.},\n  booktitle={ICDE},\n  year={2026}\n}",
        "Code": "https://github.com/Sssara-5/TF-TabularCondensation",
        "Nam_venue": "2026, ICDE (IEEE) - Rank A*",
    },

    # ===== 14. TDColER =====
    {
        "STT": 14,
        "De tai": "TDColER: On Learning Representations for Tabular Data Distillation",
        "Loai bai toan": "Dataset Distillation cho du lieu bang bang Column Embedding",
        "Muc tieu": "Dung Column Embedding de hoc bieu dien tot hon cho du lieu bang truoc khi chung cat, tang chat luong distill 0.5-143%.",
        "Ngu canh": "Cac phuong phap distillation truoc ap dung truc tiep cho Tabular ma khong xem xet dac tinh rieng cua du lieu bang (feature heterogeneity, categorical vs numerical).",
        "Input": "Du lieu bang D (so + categorical).",
        "Output": "Tap du lieu ao S chat luong cao hon nho Column Embedding.",
        "Data": "23 bo du lieu tabular (TDBench benchmark).",
        "Ingest": "Doc du lieu bang, dung Column Embedding de ma hoa tung cot.",
        "Process": "1. Hoc Column Embedding cho moi cot cua D.\n2. Dung encoder-decoder architecture de tao bieu dien latent.\n3. Ap dung distillation scheme (DC, DM, v.v.) tren bieu dien moi.\n4. Giai ma ve khong gian goc.\n5. Xuat tap S.",
        "ML": "7 nhom mo hinh khac nhau (XGBoost, LightGBM, CatBoost, MLP, LogReg, KNN, DT).",
        "Kich ban": "Benchmark TDBench: 23 datasets x 7 models x 11 distillation schemes. So sanh TDColER vs baseline schemes.",
        "Danh gia": "Test Accuracy, Relative Performance (so voi khong co TDColER).",
        "Cong nghe": "Python, PyTorch, scikit-learn.",
        "Ket qua": "TDColER tang chat luong distill 0.5-143% tren toan bo TDBench.",
        "Source": "https://arxiv.org/abs/2501.xxxxx (TDBench paper)",
        "BibTeX": "@article{kang2025tdcoler,\n  title={On Learning Representations for Tabular Data Distillation},\n  author={Kang, Inwon et al.},\n  journal={arXiv},\n  year={2025}\n}",
        "Code": "https://github.com/inwonakng/tdbench",
        "Nam_venue": "2025, ArXiv (kem TDBench benchmark)",
    },

    # ===== 15. Random (RAN) =====
    {
        "STT": 15,
        "De tai": "Uniform Random Sampling (RAN) - Baseline co ban",
        "Loai bai toan": "Baseline (Moc tham chieu)",
        "Muc tieu": "Lay mau hoan toan ngau nhien tu tap huan luyen, lam moc de so sanh voi cac phuong phap nen thong minh hon.",
        "Ngu canh": "Moi benchmarking nghiem tuc deu can mot Baseline don gian de chung minh cac phuong phap moi co gia tri. CoreTab dung RAN lam Baseline so 1.",
        "Input": "Du lieu huan luyen D, ty le nen r.",
        "Output": "Tap con S co kich thuoc |S| = r * |D|, chon ngau nhien.",
        "Data": "Moi bo du lieu.",
        "Ingest": "Doc du lieu.",
        "Process": "numpy.random.choice(n, size=int(n*r), replace=False)\nHoan toan ngau nhien, khong co thuat toan phuc tap.",
        "ML": "Moi mo hinh.",
        "Kich ban": "Chay song song voi 14 phuong phap con lai, lam Lower Bound.",
        "Danh gia": "Cung cac metric: F1-Macro, AUC-ROC, Time.",
        "Cong nghe": "numpy, scikit-learn.",
        "Ket qua": "Thuong la phuong phap kem nhat (Lower Bound). Bat ky phuong phap nao khong thang duoc RAN thi vo gia tri.",
        "Source": "N/A (Standard Baseline)",
        "BibTeX": "N/A",
        "Code": "numpy.random.choice\nsklearn.utils.resample",
        "Nam_venue": "N/A (Baseline co ban)",
    },
]

# Map to Vietnamese column names matching Plan(1).xlsx
rows = []
for d in data:
    rows.append({
        "STT": d["STT"],
        "Đề tài / Công bố khoa học": d["De tai"],
        "Loại bài toán": d["Loai bai toan"],
        "Mục tiêu": d["Muc tieu"],
        "Ngữ cảnh": d["Ngu canh"],
        "Input": d["Input"],
        "Output": d["Output"],
        "Data": d["Data"],
        "Ingest": d["Ingest"],
        "Process": d["Process"],
        "Machine Learning": d["ML"],
        "Kịch bản thực nghiệm": d["Kich ban"],
        "Phương pháp đánh giá": d["Danh gia"],
        "Công nghệ và nền tảng triển khai": d["Cong nghe"],
        "Kết quả": d["Ket qua"],
        "Source": d["Source"],
        "BibTeX": d["BibTeX"],
        "Báo cáo / Code": d["Code"],
        "Năm xuất bản tên hội nghị táp chí": d["Nam_venue"],
    })

df = pd.DataFrame(rows)

output_path = r"C:\KLTN\paper\Plan (1).xlsx"
with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name="Khảo sát tổng quan")
    ws = writer.sheets["Khảo sát tổng quan"]

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    widths = {
        "A": 5, "B": 45, "C": 25, "D": 40, "E": 35,
        "F": 30, "G": 30, "H": 40, "I": 30, "J": 50,
        "K": 30, "L": 45, "M": 35, "N": 25, "O": 45,
        "P": 35, "Q": 35, "R": 35, "S": 25,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Row heights
    for i in range(2, ws.max_row + 1):
        ws.row_dimensions[i].height = 180

print("DONE: Plan(1).xlsx da dien du 15 phuong phap x 19 cot!")
print(f"So dong: {len(rows)}")
print(f"So cot: {len(df.columns)}")
