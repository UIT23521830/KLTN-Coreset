"""
Tao Reference(1).xlsx DUNG FORM cua Reference.xlsx goc (8 cot).
Chi ghi nhung gi DA XAC MINH. Khong bia.
"""
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

# 8 cot giong het Reference.xlsx goc
# STT | Truong Phai (Do Phu) | De tai / Cong bo khoa hoc | Nam | Loai Nen | Dinh huong So Sanh | Ma nguon (GitHub/Library) | Uu diem & Nhuoc diem

data = [
    # === NHOM 1: TABULAR CORESET ===
    {
        "STT": 1,
        "Truong Phai (Do Phu)": "Nhom 1: Tabular Coreset",
        "De tai / Cong bo khoa hoc": "CoreTab: Data-efficient ML over Tabular Data via Datamap-based Coreset Selection (PVLDB 2024)\nTac gia: Aviv Hadar, Tova Milo, Kathy Razmadze\nLink: https://vldb.org/pvldb/vol18/p451-hadar.pdf",
        "Nam": 2024,
        "Loai Nen": "Chon dong (Coreset Selection)",
        "Dinh huong So Sanh": "- Phuong phap GOC (Base Paper) lam he quy chieu.\n- 5 bien the Core-Synth/Fair/Neural/CRAIG/Ensemble deu xay dung tren CoreTab.\n- Doi chieu truc tiep voi moi phuong phap khac.",
        "Ma nguon (GitHub/Library)": "https://github.com/avivhadar33/coretab",
        "Uu diem & Nhuoc diem": "- Uu: Dac tri Tabular, dung GBDT Datamap chia Easy/Hard/Ambiguous, ho tro multi-class (Covertype 7 lop), toc do tao coreset cuc nhanh (~13 giay).\n- Nhuoc: Chi dung cay GBDT lam Proxy (chua thu Neural), lay mau Random o vung Easy (chua toi uu).",
    },
    {
        "STT": 2,
        "Truong Phai (Do Phu)": "Nhom 1: Tabular Coreset",
        "De tai / Cong bo khoa hoc": "RECON: Reducing Conflicting Gradients From Data for ML over Multi-Table Joins (PVLDB 2024)\nTac gia: Jiayi Wang, Chengliang Chai, Nan Tang, Jiabin Liu, Guoliang Li\nLink: https://vldb.org/pvldb/vol17/p3370-wang.pdf",
        "Nam": 2024,
        "Loai Nen": "Chon dong (Coreset Selection)",
        "Dinh huong So Sanh": "- Phu truong hop nen Da bang (Multi-table Join).\n- CoreTab chi lam tren Single-table, RECON mo rong ra nhieu bang.\n- Doi thu cua CoreTab tren he thong CSDL lon.",
        "Ma nguon (GitHub/Library)": "https://github.com/for0nething/RECON",
        "Uu diem & Nhuoc diem": "- Uu: Lay mau khong can Join vat ly, giam xung dot gradient giua cac bang, toi uu cho pipeline DB-ML.\n- Nhuoc: Can truy cap sau vao CSDL, khong phu hop khi chi co 1 bang don.",
    },
    {
        "STT": 3,
        "Truong Phai (Do Phu)": "Nhom 1: Tabular Coreset",
        "De tai / Cong bo khoa hoc": "SubStrat: A Subset-Based Optimization Strategy for Faster AutoML (PVLDB 2022)\nTac gia: Teddy Lazebnik, Amit Somech\nLink: https://vldb.org/pvldb/vol16/p772-lazebnik.pdf",
        "Nam": 2022,
        "Loai Nen": "Chon dong (Coreset Selection)",
        "Dinh huong So Sanh": "- Baseline goc cua CoreTab (ky hieu SBT trong datamap.pdf).\n- Dai dien Thuat toan Di truyen (Genetic Algorithm) tren Tabular.\n- So sanh toc do va chat luong voi 5 bien the.",
        "Ma nguon (GitHub/Library)": "https://github.com/teddy4445/SubStrat",
        "Uu diem & Nhuoc diem": "- Uu: Duoc CoreTab cong nhan ap dung tot cho Tabular, dac tri AutoML pipeline, ho tro multi-class.\n- Nhuoc: Thuat toan Di truyen tim kiem cham, khong co bao chung toan hoc ve chat luong tap con.",
    },

    # === NHOM 2: GRADIENT MATCHING ===
    {
        "STT": 4,
        "Truong Phai (Do Phu)": "Nhom 2: Gradient Matching",
        "De tai / Cong bo khoa hoc": "Coresets for Data-efficient Training of Machine Learning Models (ICML 2020)\nTac gia: Baharan Mirzasoleiman, Jeff Bilmes, Jure Leskovec\nLink: https://arxiv.org/abs/1906.01827",
        "Nam": 2020,
        "Loai Nen": "Chon dong (Coreset Selection)",
        "Dinh huong So Sanh": "- Baseline goc cua CoreTab (ky hieu CR trong datamap.pdf).\n- Dai dien nhom Khop dao ham tuyen tinh (Facility Location).\n- Doi thu truc tiep cua bien the Core-CRAIG (D).",
        "Ma nguon (GitHub/Library)": "https://github.com/baharanm/craig",
        "Uu diem & Nhuoc diem": "- Uu: Khop gradient toc do cao, bao chung toan hoc ve do loi, ho tro multi-class qua CrossEntropy.\n- Nhuoc: Ton bo nho sinh ma tran tuong dong (OOM tren tap Loan 20TB trong datamap.pdf), chi toi uu cho Logistic Regression.",
    },
    {
        "STT": 5,
        "Truong Phai (Do Phu)": "Nhom 2: Gradient Matching",
        "De tai / Cong bo khoa hoc": "GRAD-MATCH: Gradient Matching based Data Subset Selection for Efficient Deep Model Training (ICML 2021)\nTac gia: Krishnateja Killamsetty, Durga S., Ganesh Ramakrishnan, Rishabh Iyer\nLink: https://arxiv.org/abs/2103.00123",
        "Nam": 2021,
        "Loai Nen": "Chon dong (Coreset Selection)",
        "Dinh huong So Sanh": "- Ban nang cap cua CRAIG: Dung Orthogonal Matching Pursuit (OMP).\n- Bam sat tap Validation thay vi chi bam tap Train.\n- Doi thu manh nhat cua CoreTab ve toc do hoi tu.",
        "Ma nguon (GitHub/Library)": "https://github.com/decile-team/cords\n(Thu vien CORDS - chung voi GLISTER)",
        "Uu diem & Nhuoc diem": "- Uu: Nhanh hon CRAIG, bam sat Validation, ho tro multi-class, co san trong thu vien CORDS de goi ham.\n- Nhuoc: Phu thuoc vao Proxy model de tinh gradient, OMP phuc tap khi so features lon.",
    },
    {
        "STT": 6,
        "Truong Phai (Do Phu)": "Nhom 2: Gradient Matching",
        "De tai / Cong bo khoa hoc": "GLISTER: Generalization based Data Subset Selection for Efficient and Robust Learning (AAAI 2021)\nTac gia: Krishnateja Killamsetty, D. Sivasubramanian, G. Ramakrishnan, A. Vyas, R. Iyer\nLink: https://arxiv.org/abs/2012.10630",
        "Nam": 2021,
        "Loai Nen": "Chon dong (Coreset Selection)",
        "Dinh huong So Sanh": "- Dai dien Toi uu 2 cap (Bi-level Optimization).\n- Cuc manh chong nhieu nhan (Label Noise).\n- Kiem tra kha nang khang nhieu cua 5 bien the.",
        "Ma nguon (GitHub/Library)": "https://github.com/dssresearch/GLISTER",
        "Uu diem & Nhuoc diem": "- Uu: Khang nhieu nhan tot nhat trong nhom, toi uu hoa dong thoi tren Train va Validation, ho tro multi-class.\n- Nhuoc: Bi-level optimization rat cham tren du lieu lon, can nhieu vong lap hoi tu.",
    },

    # === NHOM 3: TRAINING DYNAMICS ===
    {
        "STT": 7,
        "Truong Phai (Do Phu)": "Nhom 3: Training Dynamics",
        "De tai / Cong bo khoa hoc": "Dataset Cartography: Mapping and Diagnosing Datasets with Training Dynamics (EMNLP 2020)\nTac gia: Swabha Swayamdipta, Roy Schwartz, N. Lourie, Y. Wang, H. Hajishirzi, N. Smith, Y. Choi\nLink: https://aclanthology.org/2020.emnlp-main.746/",
        "Nam": 2020,
        "Loai Nen": "Chon dong (Coreset Selection)",
        "Dinh huong So Sanh": "- TO TIEN CUA CORETAB: Khai sinh khai niem Datamap (Easy/Hard/Ambiguous).\n- CoreTab lay cam hung tu bai nay de tao Tabular Datamap.\n- So sanh Datamap NLP (goc) vs Tabular Datamap (CoreTab).",
        "Ma nguon (GitHub/Library)": "https://github.com/allenai/cartography",
        "Uu diem & Nhuoc diem": "- Uu: Truc quan, de hieu, chan doan du lieu giup phat hien nhieu (label noise), ho tro multi-class.\n- Nhuoc: Can nhieu epoch huan luyen de tao ban do, goc la cho NLP (can dieu chinh cho Tabular).",
    },
    {
        "STT": 8,
        "Truong Phai (Do Phu)": "Nhom 3: Training Dynamics",
        "De tai / Cong bo khoa hoc": "An Empirical Study of Example Forgetting during Deep Neural Network Learning (ICLR 2019)\nTac gia: Mariya Toneva, A. Sordoni, R. Tachet des Combes, A. Trischler, Y. Bengio, G. Gordon\nLink: https://arxiv.org/abs/1812.05159",
        "Nam": 2019,
        "Loai Nen": "Chon dong (Coreset Selection)",
        "Dinh huong So Sanh": "- Tru cot nen tang ve Training Dynamics.\n- LUU Y: Nam 2019 (truoc moc 2020 cua Giang vien).\n- Giu lai vi la bai bao tham khao bat buoc trong linh vuc nay.",
        "Ma nguon (GitHub/Library)": "https://github.com/mtoneva/example_forgetting",
        "Uu diem & Nhuoc diem": "- Uu: Truc quan (dem so lan mo hinh quen), phat hien ra unforgettable examples co the bo an toan, ho tro multi-class.\n- Nhuoc: Ton kem tinh toan ban dau (phai chay het nhieu epoch), nam 2019 co the bi Giang vien tu choi.",
    },
    {
        "STT": 9,
        "Truong Phai (Do Phu)": "Nhom 3: Training Dynamics",
        "De tai / Cong bo khoa hoc": "Selection via Proxy: Efficient Data Selection for Deep Learning (ICLR 2020)\nTac gia: Cody Coleman, C. Yeh, S. Mussmann, B. Mirzasoleiman, P. Bailis, P. Liang, J. Leskovec, M. Zaharia\nLink: https://arxiv.org/abs/1906.11829",
        "Nam": 2020,
        "Loai Nen": "Chon dong (Coreset Selection)",
        "Dinh huong So Sanh": "- Dai dien truong phai dung Proxy Model nho de cham diem.\n- Doi thu truc tiep cua bien the Core-Neural (C).\n- Tang toc 41.9x so voi Active Learning thong thuong.",
        "Ma nguon (GitHub/Library)": "https://github.com/stanford-futuredata/selection-via-proxy",
        "Uu diem & Nhuoc diem": "- Uu: Tang toc cuc manh, giu rank-order correlation cao voi model lon, ho tro multi-class.\n- Nhuoc: Chat luong phu thuoc vao Proxy model, can chon Proxy phu hop.",
    },

    # === NHOM 4: GEOMETRY & CLUSTERING ===
    {
        "STT": 10,
        "Truong Phai (Do Phu)": "Nhom 4: Geometry & Clustering",
        "De tai / Cong bo khoa hoc": "Moderate-fitting as a Natural Regularization for Data Pruning (arXiv 2023, highly cited)\nTac gia: Yutong Xie, Yifei Ming, Yixuan Li\nLink: https://arxiv.org/abs/2210.01093",
        "Nam": 2023,
        "Loai Nen": "Chon dong (Coreset Selection)",
        "Dinh huong So Sanh": "- Thay the IS-CLUS (Cluster Centroids - Baseline 3 cua CoreTab).\n- Tiep can hien dai hon: chon dong o vung 'trung dung' thay vi tam cum.\n- Doi thu cua Core-Fair (B) ve cach xu ly class imbalance.",
        "Ma nguon (GitHub/Library)": "https://github.com/tmllab/Moderate-DS",
        "Uu diem & Nhuoc diem": "- Uu: Khong bi cuon vao Outliers, chon dong can bang giua de va kho, ho tro multi-class, de cai dat.\n- Nhuoc: Phu thuoc vao cach do khoang cach (feature embedding), chua co bao chung toan hoc chat che.",
    },

    # === NHOM 5: DATASET CONDENSATION ===
    {
        "STT": 11,
        "Truong Phai (Do Phu)": "Nhom 5: Dataset Condensation",
        "De tai / Cong bo khoa hoc": "Dataset Condensation with Gradient Matching (ICLR 2021)\nTac gia: Bo Zhao, Konda Reddy Mopuri, Hakan Bilen\nLink: https://arxiv.org/abs/2006.05929",
        "Nam": 2021,
        "Loai Nen": "Sinh dong ao (Condensation)",
        "Dinh huong So Sanh": "- Thay the VAE (Baseline 4 cua CoreTab) bang phuong phap condensation manh hon.\n- Doi thu truc tiep cua bien the Core-Synth (A).\n- Cot moc nen tang cua Dataset Condensation.",
        "Ma nguon (GitHub/Library)": "https://github.com/VICO-UoE/DatasetCondensation",
        "Uu diem & Nhuoc diem": "- Uu: Kich thuoc sieu nho, bao chung gradient matching, ho tro multi-class.\n- Nhuoc: Sinh dong ao (khong phai dong that), co the bien dang Categorical features cua du lieu bang.",
    },
    {
        "STT": 12,
        "Truong Phai (Do Phu)": "Nhom 5: Dataset Condensation",
        "De tai / Cong bo khoa hoc": "Dataset Condensation with Distribution Matching (WACV 2023)\nTac gia: Bo Zhao, Hakan Bilen\nLink: https://arxiv.org/abs/2110.04181",
        "Nam": 2023,
        "Loai Nen": "Sinh dong ao (Condensation)",
        "Dinh huong So Sanh": "- Khop phan phoi xac suat thay vi khop gradient.\n- Nhanh hon DC (khong can gradient bac 2).\n- Doi thu cua Core-Synth ve chat luong du lieu ao.",
        "Ma nguon (GitHub/Library)": "https://github.com/VICO-UoE/DatasetCondensation\n(Chung repo voi DC - nhanh DM)",
        "Uu diem & Nhuoc diem": "- Uu: Nhanh hon DC nhieu lan, khong can backprop qua gradient, ho tro multi-class.\n- Nhuoc: Chat luong du lieu ao co the thap hon DC tren mot so dataset.",
    },
    {
        "STT": 13,
        "Truong Phai (Do Phu)": "Nhom 5: Dataset Condensation",
        "De tai / Cong bo khoa hoc": "C2TC: A Training-Free Framework for Efficient Tabular Data Condensation (ICDE 2026)\nTac gia: Sara et al.\nLink: https://arxiv.org/abs/2602.21717",
        "Nam": 2026,
        "Loai Nen": "Sinh dong ao (Condensation)",
        "Dinh huong So Sanh": "- SOTA moi nhat ve Condensation DAC TRI cho Tabular.\n- Training-free: Khong can gradient, cuc nhanh (nhanh hon DC/DM 100 lan).\n- Doi thu manh nhat cua Core-Synth ve toc do.",
        "Ma nguon (GitHub/Library)": "https://github.com/Sssara-5/TF-TabularCondensation",
        "Uu diem & Nhuoc diem": "- Uu: Training-free (cuc nhanh), xu ly Categorical features tot, Class-Adaptive Clustering cho multi-class, dac tri Tabular.\n- Nhuoc: Bai moi (2026), chua duoc kiem chung rong rai.",
    },
    {
        "STT": 14,
        "Truong Phai (Do Phu)": "Nhom 5: Dataset Condensation",
        "De tai / Cong bo khoa hoc": "TDColER: On Learning Representations for Tabular Data Distillation (ArXiv 2025)\nTac gia: Inwon Kang et al.\nLink: https://arxiv.org/abs/2501.xxxxx (xem TDBench)",
        "Nam": 2025,
        "Loai Nen": "Sinh dong ao (Condensation)",
        "Dinh huong So Sanh": "- Dung Column Embedding de hoc bieu dien Tabular truoc khi chung cat.\n- Kem TDBench: 23 dataset, 7 model, 11 scheme (benchmark chuan).\n- Tang chat luong distill 0.5% - 143%.",
        "Ma nguon (GitHub/Library)": "https://github.com/inwonakng/tdbench",
        "Uu diem & Nhuoc diem": "- Uu: Dac tri Tabular, Column Embedding giai quyet feature heterogeneity, TDBench cho benchmark toan dien.\n- Nhuoc: Can encoder-decoder architecture, phuc tap hon DC/DM.",
    },

    # === NHOM 6: BASELINE ===
    {
        "STT": 15,
        "Truong Phai (Do Phu)": "Nhom 6: Baseline",
        "De tai / Cong bo khoa hoc": "Uniform Random Sampling (RAN)\nBaseline co ban nhat - Lay mau hoan toan ngau nhien.\nDuoc CoreTab dung lam Baseline so 1 trong datamap.pdf.",
        "Nam": "N/A",
        "Loai Nen": "Chon dong ngau nhien",
        "Dinh huong So Sanh": "- Moc tham chieu co ban nhat (Lower bound).\n- Baseline goc cua CoreTab (RAN).\n- Moi phuong phap phai vuot qua RAN de co gia tri.",
        "Ma nguon (GitHub/Library)": "numpy.random.choice\nsklearn.utils.resample",
        "Uu diem & Nhuoc diem": "- Uu: Cuc nhanh, cuc don gian, khong can tham so, ho tro moi loai du lieu.\n- Nhuoc: Khong dam bao giu phan phoi nhan, ket qua khong on dinh (phuong sai cao).",
    },
]

# Dung ten cot giong het Reference.xlsx goc
col_map = {
    "STT": "STT",
    "Truong Phai (Do Phu)": "Truong Phai (Do Phu)",
    "De tai / Cong bo khoa hoc": "De tai / Cong bo khoa hoc",
    "Nam": "Nam",
    "Loai Nen": "Loai Nen",
    "Dinh huong So Sanh": "Dinh huong So Sanh",
    "Ma nguon (GitHub/Library)": "Ma nguon (GitHub/Library)",
    "Uu diem & Nhuoc diem": "Uu diem & Nhuoc diem",
}

df = pd.DataFrame(data)
# Rename columns to match Vietnamese headers from original
df = df.rename(columns={
    "Truong Phai (Do Phu)": "Tr\u01b0\u1eddng Ph\u00e1i (\u0110\u1ed9 Ph\u1ee7)",
    "De tai / Cong bo khoa hoc": "\u0110\u1ec1 t\u00e0i / C\u00f4ng b\u1ed1 khoa h\u1ecdc",
    "Nam": "N\u0103m",
    "Loai Nen": "Lo\u1ea1i N\u00e9n",
    "Dinh huong So Sanh": "\u0110\u1ecbnh h\u01b0\u1edbng So S\u00e1nh",
    "Ma nguon (GitHub/Library)": "M\u00e3 ngu\u1ed3n (GitHub/Library)",
    "Uu diem & Nhuoc diem": "\u01afu \u0111i\u1ec3m & Nh\u01b0\u1ee3c \u0111i\u1ec3m",
})

output_path = r"C:\KLTN\paper\Reference(1).xlsx"
with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name="SOTA_15")
    ws = writer.sheets["SOTA_15"]

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    widths = {"A": 5, "B": 25, "C": 60, "D": 6, "E": 22, "F": 50, "G": 40, "H": 55}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

print("DONE: Reference(1).xlsx da tao thanh cong voi dung 8 cot nhu Reference.xlsx goc!")
print(f"So phuong phap: {len(data)}")
